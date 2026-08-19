"""Excel export helpers matching the ImportExcel Export-Excel contract used by the
PowerShell phases (bold/frozen header row, autofilter, one worksheet per dataset).

generate-dashboard.py only reads raw cell values (header row + data rows) so no
Excel "Table" styling is required for compatibility - it is added here purely for
readability when a reviewer opens the workbook directly.
"""
from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# XML 1.0 control characters openpyxl rejects (raises IllegalCharacterError), seen in
# real-world text like Defender for Endpoint vulnerability descriptions.
_ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010\013\014\016-\037]")
_MAX_CELL_LENGTH = 32767  # Excel's hard per-cell character limit


def make_output_path(prefix: str) -> str:
    """Builds a timestamped output path inside $AZWORKSHOP_OUTPUT (or the home
    directory when run standalone), mirroring each PS script's $outputFile logic.
    """
    out_dir = os.environ.get("AZWORKSHOP_OUTPUT") or os.path.expanduser("~")
    os.makedirs(out_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(out_dir, f"{prefix}_{timestamp}.xlsx")


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def add_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: List[Dict[str, Any]],
    empty_message: str = "No data found",
) -> int:
    """Adds one worksheet with a bold/frozen header row and autofilter. Returns the
    number of data rows written (0 when only the empty-state sentinel row was written,
    matching the PowerShell scripts' reporting).
    """
    reported_count = len(rows)
    if not rows:
        rows = [{"Result": empty_message}]

    columns: List[str] = []
    for row in rows:
        for key in row.keys():
            if key not in columns:
                columns.append(key)

    ws = wb.create_sheet(sheet_name[:31])
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    for row in rows:
        ws.append([_stringify(row.get(col)) for col in columns])

    sample = rows[:200]
    for i, col in enumerate(columns, start=1):
        width = max([len(str(col))] + [len(str(row.get(col, ""))) for row in sample])
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, width + 2))

    return reported_count


def save(wb: Workbook, path: str) -> None:
    wb.save(path)


def _stringify(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        value = "; ".join(str(v) for v in value)
    elif isinstance(value, dict):
        value = str(value)
    if isinstance(value, str):
        value = _ILLEGAL_CHARACTERS_RE.sub("", value)
        if len(value) > _MAX_CELL_LENGTH:
            value = value[:_MAX_CELL_LENGTH - 3] + "..."
    return value
