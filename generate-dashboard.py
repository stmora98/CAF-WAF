"""
Azure WAF/CAF Workshop - Consolidated Dashboard Generator

Deterministic Python agent that reads all Excel outputs from the launcher,
analyzes the data, generates action items, and produces a single consolidated
HTML dashboard report.

Usage:
    python3 generate-dashboard.py <output_directory>
    
    Where <output_directory> is the folder created by Launch-AzureWorkshop.ps1
    (e.g., ~/AzureWorkshop_20260730_110000)

Requirements:
    pip install openpyxl  (usually pre-installed in Cloud Shell)
"""

import sys
import os
import json
import re
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Optional
from zipfile import BadZipFile

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    os.system("pip install openpyxl --quiet")
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException


# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

class DashboardConfig:
    SEVERITY_COLORS = {
     "critical": "var(--cp-danger)",
     "high": "var(--cp-danger)",
     "medium": "var(--cp-warning)",
     "low": "var(--cp-link)",
     "info": "var(--cp-success)"
    }
    PILLAR_COLORS = {
     "Reliability": "var(--cp-link)",
     "Security": "var(--cp-danger)",
     "Cost Optimization": "var(--cp-success)",
     "Operational Excellence": "var(--cp-accent)",
     "Performance Efficiency": "var(--cp-warning)"
    }


# ═══════════════════════════════════════════════════════════════════════════════
# BILINGUAL (EN/ES) SUPPORT
# Both language variants are always rendered; CSS toggles visibility via the
# .i18n-en/.i18n-es classes based on a "lang-es" class on <body> (see the
# embedded <script> in DashboardGenerator.generate()).
# ═══════════════════════════════════════════════════════════════════════════════

def bi(en: str, es: str) -> str:
    """Wrap a piece of text in both language variants for the client-side toggle."""
    return f'<span class="i18n-en">{en}</span><span class="i18n-es">{es}</span>'


PILLAR_ES = {
    "Reliability": "Confiabilidad",
    "Security": "Seguridad",
    "Cost Optimization": "Optimización de Costos",
    "Operational Excellence": "Excelencia Operativa",
    "Performance Efficiency": "Eficiencia de Rendimiento",
}

SEVERITY_ES = {
    "critical": "CRÍTICO",
    "high": "ALTO",
    "medium": "MEDIO",
    "low": "BAJO",
    "info": "INFO",
}

SUBCATEGORY_ES = {
    "Zone Resiliency": "Resiliencia de Zona",
    "Regional Resiliency": "Resiliencia Regional",
    "Data Protection and Recovery": "Protección y Recuperación de Datos",
    "Governance and Compliance": "Gobernanza y Cumplimiento",
    "Scalability": "Escalabilidad",
    "Monitoring and Alerting": "Monitoreo y Alertas",
    "Service Upgrade and Retirement": "Actualización y Retiro de Servicios",
    "Other": "Otro",
    "Compute Optimization": "Optimización de Cómputo",
    "Storage Optimization": "Optimización de Almacenamiento",
    "Network Optimization": "Optimización de Red",
    "Data Performance": "Rendimiento de Datos",
    "Efficiency Optimization": "Optimización de Eficiencia",
    "Failure Mitigation": "Mitigación de Fallas",
    "Safe and Secure Deployment": "Implementación Segura",
}

COST_SOURCE_ES = {
    "Orphaned/unused resources": "Recursos huérfanos/no utilizados",
    "Deallocated VMs": "VMs desasignadas",
    "Idle/underutilized VMs": "VMs inactivas/subutilizadas",
    "Advisor Cost recommendations": "Recomendaciones de costos de Advisor",
}


# ═══════════════════════════════════════════════════════════════════════════════
# DATA READER
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelReader:
    """Reads Excel workbooks and returns structured data."""

    @staticmethod
    def find_workbook(base_dir: str, folder: str, prefix: str) -> Optional[str]:
        """Return the newest workbook matching a phase's output prefix."""
        phase_dir = Path(base_dir) / folder
        candidates = list(phase_dir.glob(f"{prefix}*.xlsx"))
        if not candidates:
            print(f"  ! {prefix}: no workbook found in {phase_dir}")
            return None
        return str(max(candidates, key=lambda path: path.stat().st_mtime))
    
    @staticmethod
    def read_workbook(path: str) -> dict:
        """Read all sheets from an Excel file into a dict of lists-of-dicts."""
        if not os.path.exists(path):
            return {}
        
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError) as exc:
            print(f"  ! Skipping invalid workbook {path}: {exc}")
            return {}
        result = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                result[sheet_name] = []
                continue
            
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:]:
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = val if val is not None else ""
                data.append(record)
            result[sheet_name] = data
        
        wb.close()
        return result
    
    @staticmethod
    def read_all(base_dir: str) -> dict:
        """Read all Excel files from the workshop output directory."""
        data = {
            "discovery": {},
            "advisor": {},
            "metrics": {},
            "governance": {},
            "security": {},
            "checklists": {}
        }
        
        # Discovery
        path = ExcelReader.find_workbook(base_dir, "01_Discovery", "AzureDiscovery")
        if path:
            data["discovery"] = ExcelReader.read_workbook(path)
            print(f"  ✓ Discovery: {sum(len(v) for v in data['discovery'].values())} records across {len(data['discovery'])} sheets")
        
        # Advisor
        path = ExcelReader.find_workbook(base_dir, "02_Advisor", "AzureAdvisor")
        if path:
            data["advisor"] = ExcelReader.read_workbook(path)
            print(f"  ✓ Advisor: {sum(len(v) for v in data['advisor'].values())} records across {len(data['advisor'])} sheets")
        
        # Metrics
        path = ExcelReader.find_workbook(base_dir, "03_Metrics", "AzureMetrics")
        if path:
            data["metrics"] = ExcelReader.read_workbook(path)
            print(f"  ✓ Metrics: {sum(len(v) for v in data['metrics'].values())} records across {len(data['metrics'])} sheets")
        
        # Governance
        path = ExcelReader.find_workbook(base_dir, "04_Governance", "AzureGovernance")
        if path:
            data["governance"] = ExcelReader.read_workbook(path)
            print(f"  ✓ Governance: {sum(len(v) for v in data['governance'].values())} records across {len(data['governance'])} sheets")

        # Security
        path = ExcelReader.find_workbook(base_dir, "05_Security", "AzureSecurity")
        if path:
            data["security"] = ExcelReader.read_workbook(path)
            print(f"  ✓ Security: {sum(len(v) for v in data['security'].values())} records across {len(data['security'])} sheets")

        # Checklists (Azure/review-checklists community WAF ARG compliance scan)
        path = ExcelReader.find_workbook(base_dir, "06_Checklists", "AzureChecklists")
        if path:
            data["checklists"] = ExcelReader.read_workbook(path)
            print(f"  ✓ Checklists: {sum(len(v) for v in data['checklists'].values())} records across {len(data['checklists'])} sheets")
        
        return data


# ═══════════════════════════════════════════════════════════════════════════════
# ADVISOR SCORE MODEL
# ═══════════════════════════════════════════════════════════════════════════════

class AdvisorScoreModel:
    """
    Implements the official Azure Advisor Score methodology using the data this
    toolkit collects.
    Reference: https://learn.microsoft.com/en-us/azure/advisor/advisor-score#calculation-of-advisor-score

    Official formula (Reliability / Performance / Operational Excellence):
        Subcategory Score = (Healthy Resources / Total Applicable Resources) * 100
        Category Score = sum((Healthy/Total) * SubcategoryWeight) / sum(SubcategoryWeight) * 100

    Deviations from the live Advisor score (documented, not hidden):
      - Subcategory weights below are copied verbatim from the Microsoft Learn doc.
      - Recommendations are mapped to subcategories via keyword matching on the
        recommendation problem/solution text, because Azure Resource Graph does not
        expose Advisor's internal subcategory tag.
      - "Total Applicable Resources" uses the total discovered resource inventory
        (01_Discovery/ResourceSummary) as a shared pool, because Advisor's exact
        per-subcategory assessed-resource counts aren't exposed outside the portal.
      - Security score uses the Microsoft Defender Secure Score directly (the same
        model Advisor uses for this category), read from 04_Governance/SecureScores.
      - Cost score uses a resource-count healthy ratio instead of retail-cost
        weighting, because this toolkit doesn't call the Azure Retail Prices API.
    """

    RELIABILITY_WEIGHTS = {
        "Zone Resiliency": 30,
        "Regional Resiliency": 25,
        "Data Protection and Recovery": 20,
        "Governance and Compliance": 10,
        "Scalability": 10,
        "Monitoring and Alerting": 5,
        "Service Upgrade and Retirement": 5,
        "Other": 5,
    }
    PERFORMANCE_WEIGHTS = {
        "Compute Optimization": 25,
        "Storage Optimization": 25,
        "Network Optimization": 25,
        "Data Performance": 20,
        "Scalability": 10,
        "Monitoring and Alerting": 5,
        "Service Upgrade and Retirement": 5,
        "Other": 5,
    }
    OPERATIONAL_WEIGHTS = {
        "Efficiency Optimization": 30,
        "Failure Mitigation": 20,
        "Scalability": 10,
        "Monitoring and Alerting": 5,
        "Safe and Secure Deployment": 5,
        "Service Upgrade and Retirement": 5,
        "Other": 5,
    }

    RELIABILITY_RULES = [
        ("Zone Resiliency", ["availability zone", "zone redundant", "zone-redundant", "zonal"]),
        ("Regional Resiliency", ["paired region", "geo-redundant", "geo redundant", "multi-region", "georeplication", "geo replication", "ddos"]),
        ("Data Protection and Recovery", ["backup", "recovery vault", "restore", "snapshot", "soft delete", "point-in-time"]),
        ("Governance and Compliance", ["policy", "compliance", "governance"]),
        ("Scalability", ["scale", "autoscale", "throughput", "capacity"]),
        ("Monitoring and Alerting", ["diagnostic", "alert", "monitor", "log analytics", "insights"]),
        ("Service Upgrade and Retirement", ["upgrade", "retire", "deprecat", "migrate", "end of life", "eol", "outdated version"]),
    ]
    PERFORMANCE_RULES = [
        ("Compute Optimization", ["virtual machine", "vm", "app service plan", "compute"]),
        ("Storage Optimization", ["storage account", "disk", "data warehouse", "blob"]),
        ("Network Optimization", ["traffic manager", "network", "express route", "vpn", "load balancer", "cdn", "front door"]),
        ("Data Performance", ["sql", "database", "cosmos", "query"]),
        ("Scalability", ["scale", "autoscale", "throughput", "capacity"]),
        ("Monitoring and Alerting", ["diagnostic", "alert", "monitor", "log analytics", "insights"]),
        ("Service Upgrade and Retirement", ["upgrade", "retire", "deprecat", "migrate", "end of life", "eol", "outdated version"]),
    ]
    OPERATIONAL_RULES = [
        ("Efficiency Optimization", ["accelerated networking", "efficiency", "configuration"]),
        ("Failure Mitigation", ["failure", "deployment failure", "resiliency", "mitigat"]),
        ("Safe and Secure Deployment", ["safe deployment", "secure deployment", "rollout", "blue-green", "canary", "staged rollout"]),
        ("Scalability", ["scale", "autoscale", "throughput", "capacity"]),
        ("Monitoring and Alerting", ["diagnostic", "alert", "monitor", "log analytics", "insights"]),
        ("Service Upgrade and Retirement", ["upgrade", "retire", "deprecat", "migrate", "end of life", "eol", "outdated version"]),
    ]

    @staticmethod
    def classify(text: str, rules) -> str:
        t = (text or "").lower()
        for subcategory, keywords in rules:
            for kw in keywords:
                if re.search(r"\b" + re.escape(kw.strip()) + r"\b", t):
                    return subcategory
        return "Other"

    @staticmethod
    def bucket(records, rules) -> dict:
        """Group impacted-resource keys into subcategories via keyword classification."""
        buckets = {}
        for rec in records:
            text = " ".join(str(rec.get(f, "")) for f in ("problem", "solution", "impactedType", "title"))
            subcategory = AdvisorScoreModel.classify(text, rules)
            key = str(rec.get("impactedResource") or rec.get("name") or rec.get("Name") or id(rec))
            buckets.setdefault(subcategory, set()).add(key)
        return buckets

    @staticmethod
    def merge_buckets(*bucket_dicts) -> dict:
        merged = {}
        for buckets in bucket_dicts:
            for subcategory, keys in buckets.items():
                merged.setdefault(subcategory, set()).update(keys)
        return merged

    @staticmethod
    def score_category(weights: dict, buckets: dict, total_pool: int, forced_zero: set = frozenset()):
        """Official formula: sum((Healthy/Total)*Weight) / sum(Weight) * 100."""
        if total_pool <= 0:
            return None, []
        rows = []
        weighted_sum = 0.0
        weight_total = 0
        for subcategory, weight in weights.items():
            if subcategory in forced_zero:
                impacted = total_pool
            else:
                impacted = min(len(buckets.get(subcategory, set())), total_pool)
            healthy = max(0, total_pool - impacted)
            pct = (healthy / total_pool) * 100
            weighted_sum += pct * weight
            weight_total += weight
            rows.append({
                "subcategory": subcategory, "weight": weight,
                "healthy": healthy, "total": total_pool, "pct": round(pct, 1)
            })
        score = weighted_sum / weight_total if weight_total else None
        return (round(score) if score is not None else None), rows


# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class InsightEngine:
    """Analyzes collected data, scores each WAF pillar, and generates action items."""
    
    def __init__(self, data: dict):
        self.data = data
        self.insights = []
        self.action_items = []
        self.scores = {
            "Reliability": 0,
            "Security": 0,
            "Cost Optimization": 0,
            "Operational Excellence": 0,
            "Performance Efficiency": 0
        }
        self.breakdowns = {}
        self.max_score = 100

    def _resource_pool(self) -> int:
        """Total discovered resources, used as the shared 'assessed resource' pool."""
        summary = self.data.get("discovery", {}).get("ResourceSummary", [])
        return sum(int(r.get("Count", 0) or 0) for r in summary)

    @staticmethod
    def _data_rows(rows: list) -> list:
        """Exclude sentinel rows emitted when a workbook query returns no data."""
        empty_results = {"no data found", "no resources found"}
        return [
            row for row in rows
            if str(row.get("Result", row.get("result", ""))).strip().lower() not in empty_results
            and str(row.get("DataStatus", "")).strip().lower() != "nodata"
        ]
    
    def analyze(self):
        """Run all analysis passes."""
        self._analyze_reliability()
        self._analyze_security()
        self._analyze_cost()
        self._analyze_operational_excellence()
        self._analyze_performance()
        self._analyze_checklists()
        return self

    def _analyze_checklists(self):
        """Community WAF checks from Azure/review-checklists, run via Resource Graph.
        Adds action items only — pillar scores above still follow the Advisor Score model.
        Source: https://github.com/Azure/review-checklists
        """
        rows = self._data_rows(self.data.get("checklists", {}).get("Findings", []))
        severity_map = {"high": "high", "medium": "medium", "low": "low"}
        grouped = {}
        for row in rows:
            key = row.get("Guid") or row.get("Text")
            if not key:
                continue
            bucket = grouped.setdefault(key, {
                "pillar": row.get("WafPillar") or "Operational Excellence",
                "severity": severity_map.get(str(row.get("Severity", "")).strip().lower(), "medium"),
                "service": str(row.get("Service", "")),
                "text": str(row.get("Text", "")),
                "link": str(row.get("Link", "")),
                "resources": []
            })
            resource_id = row.get("ResourceId", "")
            if resource_id:
                bucket["resources"].append(escape(str(resource_id)))

        for bucket in grouped.values():
            if bucket["pillar"] not in self.scores:
                continue
            text = escape(bucket["text"])
            title = text if len(text) <= 90 else text[:87] + "..."
            link_html = f' <a href="{escape(bucket["link"])}" target="_blank">Learn more</a>' if bucket["link"] else ""
            count = len(bucket["resources"])
            desc = (f'{text}{link_html} — {count} non-compliant resource(s). '
                    f'Source: Azure/review-checklists community WAF checklist ({escape(bucket["service"])}).')
            self.action_items.append({
                "pillar": bucket["pillar"],
                "severity": bucket["severity"],
                "title": f'[{escape(bucket["service"])}] {title}',
                "title_es": f'[{escape(bucket["service"])}] {title}',
                "description": desc,
                "description_es": desc,
                "resources": bucket["resources"][:10]
            })
    
    def _analyze_reliability(self):
        """Reliability score using the official Advisor Score subcategory model."""
        total_pool = self._resource_pool()
        forced_zero = set()
        synthetic = []

        # Check VMs without availability zones
        vms = self.data.get("discovery", {}).get("VMs", [])
        vms_no_zone = [v for v in vms if not v.get("zone") and not v.get("availabilityZone")]
        synthetic += [{"problem": "Virtual machine not deployed to an Availability Zone",
                       "impactedResource": v.get("name", v.get("Name", ""))} for v in vms_no_zone]
        if vms_no_zone:
            self.action_items.append({
                "pillar": "Reliability",
                "severity": "high",
                "title": f"{len(vms_no_zone)} VMs without Availability Zones",
                "title_es": f"{len(vms_no_zone)} VMs sin Zonas de Disponibilidad",
                "description": "These VMs are not protected against datacenter failures. Consider migrating to zone-redundant deployments.",
                "description_es": "Estas VMs no están protegidas contra fallas del centro de datos. Considere migrar a implementaciones con redundancia de zona.",
                "resources": [v.get("name", v.get("Name", "")) for v in vms_no_zone[:10]]
            })
        
        # Check backup vaults
        backups = self.data.get("discovery", {}).get("BackupVaults", [])
        if not backups or (len(backups) == 1 and backups[0].get("Result") == "No resources found"):
            forced_zero.add("Data Protection and Recovery")
            self.action_items.append({
                "pillar": "Reliability",
                "severity": "critical",
                "title": "No Backup/Recovery Services Vaults found",
                "title_es": "No se encontraron Vaults de Backup/Recovery Services",
                "description": "No backup infrastructure detected. Critical workloads should have backup policies configured.",
                "description_es": "No se detectó infraestructura de copia de seguridad. Las cargas de trabajo críticas deben tener políticas de backup configuradas.",
                "resources": []
            })
        
        # Check DDoS protection on VNets
        vnets = self.data.get("discovery", {}).get("VNets", []) or self.data.get("governance", {}).get("VNets", [])
        vnets_no_ddos = [v for v in vnets if str(v.get("ddos", v.get("ddosProtection", v.get("enableDdosProtection", "")))).lower() in ("false", "", "none")]
        synthetic += [{"problem": "VNet without DDoS Protection Standard",
                       "impactedResource": v.get("name", v.get("Name", ""))} for v in vnets_no_ddos]
        if vnets_no_ddos:
            self.action_items.append({
                "pillar": "Reliability",
                "severity": "medium",
                "title": f"{len(vnets_no_ddos)} VNets without DDoS Protection",
                "title_es": f"{len(vnets_no_ddos)} VNets sin Protección DDoS",
                "description": "Enable Azure DDoS Protection Standard for internet-facing workloads.",
                "description_es": "Habilite Azure DDoS Protection Standard para cargas de trabajo expuestas a Internet.",
                "resources": [v.get("name", v.get("Name", "")) for v in vnets_no_ddos[:10]]
            })
        
        # Advisor HighAvailability recommendations
        ha_recs = self.data.get("advisor", {}).get("Reliability", [])
        high_impact_ha = [r for r in ha_recs if str(r.get("impact", "")).lower() == "high"]
        if high_impact_ha:
            self.action_items.append({
                "pillar": "Reliability",
                "severity": "high",
                "title": f"{len(high_impact_ha)} High-Impact Reliability recommendations from Advisor",
                "title_es": f"{len(high_impact_ha)} Recomendaciones de Confiabilidad de alto impacto de Advisor",
                "description": "Azure Advisor has identified critical reliability improvements.",
                "description_es": "Azure Advisor identificó mejoras críticas de confiabilidad.",
                "resources": [r.get("impactedResource", r.get("impactedValue", "")) for r in high_impact_ha[:10]]
            })

        buckets = AdvisorScoreModel.merge_buckets(
            AdvisorScoreModel.bucket(ha_recs, AdvisorScoreModel.RELIABILITY_RULES),
            AdvisorScoreModel.bucket(synthetic, AdvisorScoreModel.RELIABILITY_RULES),
        )
        score, rows = AdvisorScoreModel.score_category(
            AdvisorScoreModel.RELIABILITY_WEIGHTS, buckets, total_pool, forced_zero)
        self.scores["Reliability"] = score if score is not None else 0
        self.breakdowns["Reliability"] = {"available": score is not None, "rows": rows, "pool": total_pool}
    
    def _analyze_security(self):
        """Security score = Microsoft Defender Secure Score average (same model Advisor uses)."""
        security = self.data.get("security", {})
        secure_scores = self._data_rows(security.get("SecureScores", []))
        if not secure_scores:
            secure_scores = self._data_rows(self.data.get("governance", {}).get("SecureScores", []))
        avg_score = None
        if secure_scores:
            values = [
                float(s.get("percentageScore", s.get("pct", s.get("percentage", 0))) or 0)
                for s in secure_scores
            ]
            avg_score = sum(values) / len(values)
            if avg_score <= 1:  # export may store a 0-1 ratio instead of 0-100
                avg_score *= 100
            self.insights.append({
                "pillar": "Security",
                "text": f"Average Defender Secure Score: {avg_score:.0f}%"
            })
        
        # Storage accounts with public access
        storage = self.data.get("discovery", {}).get("Storage", [])
        public_storage = [s for s in storage if str(s.get("publicBlob", s.get("allowBlobPublicAccess", ""))).lower() == "true"]
        if public_storage:
            self.action_items.append({
                "pillar": "Security",
                "severity": "critical",
                "title": f"{len(public_storage)} Storage accounts allow public blob access",
                "title_es": f"{len(public_storage)} Cuentas de almacenamiento permiten acceso público a blobs",
                "description": "Disable public blob access unless explicitly required. This is a common data exposure risk.",
                "description_es": "Deshabilite el acceso público a blobs a menos que sea explícitamente necesario. Este es un riesgo común de exposición de datos.",
                "resources": [s.get("name", s.get("Name", "")) for s in public_storage[:10]]
            })
        
        # Key Vaults without purge protection
        keyvaults = self.data.get("discovery", {}).get("KeyVaults", []) or self.data.get("governance", {}).get("KeyVaults", [])
        kv_no_purge = [k for k in keyvaults if str(k.get("purgeProtection", k.get("enablePurgeProtection", ""))).lower() in ("false", "", "none")]
        if kv_no_purge:
            self.action_items.append({
                "pillar": "Security",
                "severity": "high",
                "title": f"{len(kv_no_purge)} Key Vaults without Purge Protection",
                "title_es": f"{len(kv_no_purge)} Key Vaults sin Protección contra Purga",
                "description": "Enable purge protection to prevent permanent deletion of secrets, keys, and certificates.",
                "description_es": "Habilite la protección contra purga para evitar la eliminación permanente de secretos, claves y certificados.",
                "resources": [k.get("name", k.get("Name", "")) for k in kv_no_purge[:10]]
            })
        
        # Storage without TLS 1.2
        old_tls = [s for s in storage if s.get("tlsVersion", s.get("minimumTlsVersion", "")) and "1.2" not in str(s.get("tlsVersion", s.get("minimumTlsVersion", "")))]
        if old_tls:
            self.action_items.append({
                "pillar": "Security",
                "severity": "high",
                "title": f"{len(old_tls)} Storage accounts not enforcing TLS 1.2",
                "title_es": f"{len(old_tls)} Cuentas de almacenamiento sin exigir TLS 1.2",
                "description": "Enforce minimum TLS 1.2 for all storage accounts to prevent protocol downgrade attacks.",
                "description_es": "Exija TLS 1.2 como mínimo en todas las cuentas de almacenamiento para evitar ataques de downgrade de protocolo.",
                "resources": [s.get("name", s.get("Name", "")) for s in old_tls[:10]]
            })
        
        # Security advisor recommendations (informational — already reflected in Secure Score)
        sec_recs = self._data_rows(self.data.get("advisor", {}).get("Security", []))
        if sec_recs:
            self.action_items.append({
                "pillar": "Security",
                "severity": "medium",
                "title": f"{len(sec_recs)} Security recommendations from Advisor",
                "title_es": f"{len(sec_recs)} Recomendaciones de seguridad de Advisor",
                "description": "Review Azure Advisor security findings and remediate.",
                "description_es": "Revise los hallazgos de seguridad de Azure Advisor y corríjalos.",
                "resources": [r.get("impactedResource", r.get("impactedValue", "")) for r in sec_recs[:10]]
            })
        
        # Broad-scope Owner/Contributor role assignments (RBAC hygiene, informational)
        OWNER_ROLE_ID = "8e3af657-a8ff-443c-a75c-2fe8c4bcb635"
        CONTRIBUTOR_ROLE_ID = "b24988ac-6180-42a0-ab88-20f7382dd24c"
        role_assignments = self.data.get("governance", {}).get("RoleAssignments", [])
        broad_privileged = [
            r for r in role_assignments
            if any(rid in str(r.get("roleDefId", "")) for rid in (OWNER_ROLE_ID, CONTRIBUTOR_ROLE_ID))
            and "resourceGroups" not in str(r.get("scope_", ""))
        ]
        if broad_privileged:
            self.action_items.append({
                "pillar": "Security",
                "severity": "medium",
                "title": f"{len(broad_privileged)} Owner/Contributor role assignments at subscription scope or higher",
                "title_es": f"{len(broad_privileged)} Asignaciones de rol Owner/Contributor con alcance de suscripción o superior",
                "description": "Review broad-scope privileged role assignments for least-privilege access. Prefer scoping to resource groups.",
                "description_es": "Revise las asignaciones de roles privilegiados de alcance amplio para aplicar el principio de mínimo privilegio. Prefiera limitar el alcance a grupos de recursos.",
                "resources": [f"{r.get('principalType', '')} ({r.get('scope_', '')})" for r in broad_privileged[:10]]
            })
        
        # Microsoft Defender for Cloud plans not enabled (informational)
        defender_plans = self._data_rows(security.get("DefenderPlans", []))
        if not defender_plans:
            defender_plans = self._data_rows(self.data.get("governance", {}).get("DefenderPlans", []))
        disabled_plans = [
            p for p in defender_plans
            if str(p.get("pricingTier", p.get("tier", ""))).lower() == "free"
        ]
        if disabled_plans:
            self.action_items.append({
                "pillar": "Security",
                "severity": "medium",
                "title": f"{len(disabled_plans)} Microsoft Defender for Cloud plans not enabled (Free tier)",
                "title_es": f"{len(disabled_plans)} Planes de Microsoft Defender for Cloud no habilitados (nivel gratuito)",
                "description": "Enable Microsoft Defender for Cloud plans for full threat protection coverage.",
                "description_es": "Habilite los planes de Microsoft Defender for Cloud para una cobertura completa de protección contra amenazas.",
                "resources": sorted({p.get("planName", p.get("name", "")) for p in disabled_plans})[:10]
            })

        posture_recommendations = self._data_rows(security.get("Recommendations", []))
        open_recommendations = [
            recommendation for recommendation in posture_recommendations
            if str(recommendation.get("recommendationState", "")).lower() not in ("healthy", "notapplicable")
        ]
        severe_recommendations = [
            recommendation for recommendation in open_recommendations
            if str(recommendation.get("recommendationSeverity", "")).lower() in ("critical", "high")
        ]
        if severe_recommendations:
            self.action_items.append({
                "pillar": "Security",
                "severity": "high",
                "title": f"{len(severe_recommendations)} high-severity Defender for Cloud recommendations",
                "title_es": f"{len(severe_recommendations)} recomendaciones de alta severidad de Defender for Cloud",
                "description": "Prioritize unhealthy CSPM recommendations that carry critical or high severity.",
                "description_es": "Priorice las recomendaciones de CSPM no saludables con severidad crítica o alta.",
                "resources": [
                    recommendation.get("recommendationName", recommendation.get("affectedResourceId", ""))
                    for recommendation in severe_recommendations[:10]
                ]
            })

        score_controls = self._data_rows(security.get("ScoreControls", []))
        unhealthy_control_resources = sum(
            int(control.get("unhealthyResourceCount", 0) or 0) for control in score_controls
        )
        mcsb_assessments = self._data_rows(security.get("MCSBCompliance", []))
        mcsb_failed_resources = sum(
            int(assessment.get("failedResources", 0) or 0) for assessment in mcsb_assessments
        )

        incidents = self._data_rows(security.get("Incidents", []))
        active_incidents = [
            incident for incident in incidents
            if str(incident.get("Status", "")).lower() not in ("resolved", "redirected")
        ]
        severe_incidents = [
            incident for incident in active_incidents
            if str(incident.get("Severity", "")).lower() in ("critical", "high")
        ]
        if severe_incidents:
            self.action_items.append({
                "pillar": "Security",
                "severity": "critical",
                "title": f"{len(severe_incidents)} active high-severity security incidents",
                "title_es": f"{len(severe_incidents)} incidentes de seguridad activos de alta severidad",
                "description": "Investigate active Microsoft Defender XDR incidents and confirm ownership and containment.",
                "description_es": "Investigue los incidentes activos de Microsoft Defender XDR y confirme su asignación y contención.",
                "resources": [incident.get("DisplayName", incident.get("IncidentId", "")) for incident in severe_incidents[:10]]
            })

        graph_alerts = self._data_rows(security.get("Alerts", []))
        cloud_alerts = self._data_rows(security.get("CloudAlerts", []))
        active_alerts = [
            alert for alert in graph_alerts
            if str(alert.get("Status", "")).lower() not in ("resolved", "dismissed")
        ]
        active_cloud_alerts = [
            alert for alert in cloud_alerts
            if str(alert.get("status", "")).lower() not in ("resolved", "dismissed")
        ]
        source_status = self._data_rows(security.get("SourceStatus", []))
        
        self.scores["Security"] = round(avg_score) if avg_score is not None else 0
        self.breakdowns["Security"] = {
            "available": avg_score is not None,
            "secure_score_pct": round(avg_score, 1) if avg_score is not None else None,
            "subscriptions_scored": len(secure_scores),
            "findings": [
                ("Public Blob Access on Storage Accounts", "Acceso p\u00fablico a blobs en cuentas de almacenamiento", len(public_storage)),
                ("Key Vaults without Purge Protection", "Key Vaults sin Protecci\u00f3n contra Purga", len(kv_no_purge)),
                ("Storage Accounts below TLS 1.2", "Cuentas de almacenamiento por debajo de TLS 1.2", len(old_tls)),
                ("Broad-scope Owner/Contributor Assignments", "Asignaciones Owner/Contributor de alcance amplio", len(broad_privileged)),
                ("Defender for Cloud Plans on Free Tier", "Planes de Defender for Cloud en nivel gratuito", len(disabled_plans)),
                ("Open Defender for Cloud Recommendations", "Recomendaciones abiertas de Defender for Cloud", len(open_recommendations)),
                ("Unhealthy Secure Score Control Resources", "Recursos no saludables en controles de Secure Score", unhealthy_control_resources),
                ("Failed MCSB Resources", "Recursos con error en MCSB", mcsb_failed_resources),
                ("Active Defender XDR Incidents", "Incidentes activos de Defender XDR", len(active_incidents)),
                ("Active Security Alerts", "Alertas de seguridad activas", len(active_alerts) + len(active_cloud_alerts)),
                ("Advisor Security Recommendations", "Recomendaciones de seguridad de Advisor", len(sec_recs)),
            ],
            "open_recommendations": len(open_recommendations),
            "severe_recommendations": len(severe_recommendations),
            "unhealthy_control_resources": unhealthy_control_resources,
            "mcsb_assessments": len(mcsb_assessments),
            "mcsb_failed_resources": mcsb_failed_resources,
            "defender_plans": len(defender_plans),
            "disabled_plans": len(disabled_plans),
            "active_incidents": len(active_incidents),
            "severe_incidents": len(severe_incidents),
            "active_alerts": len(active_alerts) + len(active_cloud_alerts),
            "source_status": source_status,
        }
    
    def _analyze_cost(self):
        """Cost score = healthy-resource ratio (count-based proxy for Advisor's retail-cost weighting)."""
        total_pool = self._resource_pool()
        impacted_keys = set()
        sources = []
        
        # Orphaned resources
        orphaned = self._data_rows(self.data.get("governance", {}).get("OrphanedResources", []))
        if not orphaned:
            # Try discovery sheets
            disks = self._data_rows(self.data.get("discovery", {}).get("UnattachedDisks", []))
            orphaned = disks
        
        if orphaned:
            keys = {str(o.get("Name", o.get("name", id(o)))) for o in orphaned}
            impacted_keys |= keys
            sources.append({
                "name": "Orphaned/unused resources",
                "count": len(keys),
                "unit": "resources",
                "affected": len(keys),
            })
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "medium",
                "title": f"{len(orphaned)} Orphaned/unused resources detected",
                "title_es": f"{len(orphaned)} Recursos huérfanos/no utilizados detectados",
                "description": "These resources incur costs but are not attached to any workload. Review and delete if unused.",
                "description_es": "Estos recursos generan costos pero no están asociados a ninguna carga de trabajo. Revíselos y elimínelos si no se utilizan.",
                "resources": [o.get("Name", o.get("name", "")) for o in orphaned[:10]]
            })
        
        # VMs that are deallocated (still paying for disks)
        dealloc_vms = self._data_rows(self.data.get("discovery", {}).get("DeallocatedVMs", []))
        if dealloc_vms:
            keys = {str(v.get("name", v.get("Name", id(v)))) for v in dealloc_vms}
            impacted_keys |= keys
            sources.append({
                "name": "Deallocated VMs",
                "count": len(keys),
                "unit": "resources",
                "affected": len(keys),
            })
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "low",
                "title": f"{len(dealloc_vms)} Deallocated VMs (still paying for disks/IPs)",
                "title_es": f"{len(dealloc_vms)} VMs desasignadas (aún generan costo por discos/IPs)",
                "description": "Deallocated VMs still incur costs for attached disks and static IPs. Consider deleting if no longer needed.",
                "description_es": "Las VMs desasignadas siguen generando costos por los discos e IPs estáticas asociadas. Considere eliminarlas si ya no son necesarias.",
                "resources": [v.get("name", v.get("Name", "")) for v in dealloc_vms[:10]]
            })
        
        # Right-sizing from metrics
        vm_metrics = self._data_rows(self.data.get("metrics", {}).get("VM_RightSizing", []))
        underutilized = [v for v in vm_metrics if "Idle" in str(v.get("Assessment", "")) or "Underutilized" in str(v.get("Assessment", ""))]
        if underutilized:
            keys = {str(v.get("Name", id(v))) for v in underutilized}
            impacted_keys |= keys
            sources.append({
                "name": "Idle/underutilized VMs",
                "count": len(keys),
                "unit": "resources",
                "affected": len(keys),
            })
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "high",
                "title": f"{len(underutilized)} VMs are idle or underutilized",
                "title_es": f"{len(underutilized)} VMs están inactivas o subutilizadas",
                "description": "These VMs have <15% average CPU over 30 days. Consider downsizing or deallocating.",
                "description_es": "Estas VMs tienen <15% de uso promedio de CPU en 30 días. Considere reducir su tamaño o desasignarlas.",
                "resources": [f"{v.get('Name', '')} ({v.get('VMSize', '')} @ {v.get('AvgCPU_Pct', '?')}% CPU)" for v in underutilized[:10]]
            })
        
        # SQL databases that are oversized or underutilized (DTU/CPU-based)
        sql_metrics = self._data_rows(self.data.get("metrics", {}).get("SQL_RightSizing", []))
        sql_waste = [d for d in sql_metrics if str(d.get("Assessment", "")).startswith(("Oversized", "Underutilized"))]
        if sql_waste:
            keys = {str(d.get("Name", id(d))) for d in sql_waste}
            impacted_keys |= keys
            sources.append({
                "name": "Oversized/underutilized SQL databases",
                "count": len(keys),
                "unit": "resources",
                "affected": len(keys),
            })
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "medium",
                "title": f"{len(sql_waste)} SQL databases are oversized or underutilized",
                "title_es": f"{len(sql_waste)} bases de datos SQL están sobredimensionadas o subutilizadas",
                "description": "These databases run well below their provisioned DTU/CPU tier over 30 days. Consider a lower SKU.",
                "description_es": "Estas bases de datos operan muy por debajo de su nivel de DTU/CPU aprovisionado durante 30 días. Considere un SKU menor.",
                "resources": [f"{d.get('Name', '')} ({d.get('SKU', '')} @ {d.get('AvgUsage_Pct', '?')}% {d.get('MetricType', '')})" for d in sql_waste[:10]]
            })

        # App Service Plans that are idle or underutilized
        plan_metrics = self._data_rows(self.data.get("metrics", {}).get("AppPlan_RightSizing", []))
        plan_waste = [p for p in plan_metrics if str(p.get("Assessment", "")).startswith(("Idle", "Underutilized"))]
        if plan_waste:
            keys = {str(p.get("Name", id(p))) for p in plan_waste}
            impacted_keys |= keys
            sources.append({
                "name": "Idle/underutilized App Service Plans",
                "count": len(keys),
                "unit": "resources",
                "affected": len(keys),
            })
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "medium",
                "title": f"{len(plan_waste)} App Service Plans are idle or underutilized",
                "title_es": f"{len(plan_waste)} planes de App Service están inactivos o subutilizados",
                "description": "These plans run well below their provisioned CPU/memory over 30 days. Consider scaling down or consolidating apps.",
                "description_es": "Estos planes operan muy por debajo de su CPU/memoria aprovisionada durante 30 días. Considere reducir su tamaño o consolidar aplicaciones.",
                "resources": [f"{p.get('Name', '')} ({p.get('SKU', '')} @ {p.get('AvgCPU_Pct', '?')}% CPU)" for p in plan_waste[:10]]
            })

        # Storage accounts with zero/minimal activity
        storage_metrics = self._data_rows(self.data.get("metrics", {}).get("Storage_Activity", []))
        storage_waste = [s for s in storage_metrics if str(s.get("Assessment", "")).startswith(("Zero", "Minimal"))]
        if storage_waste:
            keys = {str(s.get("Name", id(s))) for s in storage_waste}
            impacted_keys |= keys
            sources.append({
                "name": "Storage accounts with zero/minimal activity",
                "count": len(keys),
                "unit": "resources",
                "affected": len(keys),
            })
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "low",
                "title": f"{len(storage_waste)} Storage accounts show zero/minimal activity",
                "title_es": f"{len(storage_waste)} cuentas de almacenamiento muestran actividad nula/mínima",
                "description": "These storage accounts had near-zero transactions over 30 days. Review for deletion or tier downgrade (e.g. to Archive).",
                "description_es": "Estas cuentas de almacenamiento tuvieron transacciones casi nulas durante 30 días. Considere eliminarlas o cambiarlas a un nivel más económico (p. ej. Archive).",
                "resources": [f"{s.get('Name', '')} ({s.get('SKU', '')} @ {s.get('AvgDailyTxns', '?')} txns/day)" for s in storage_waste[:10]]
            })

        # Cost recommendations from Advisor
        cost_recs = self._data_rows(self.data.get("advisor", {}).get("Cost", []))
        if cost_recs:
            keys = {str(r.get("impactedResource", r.get("impactedValue", id(r)))) for r in cost_recs}
            impacted_keys |= keys
            sources.append({
                "name": "Advisor Cost recommendations",
                "count": len(cost_recs),
                "unit": "recommendations",
                "affected": len(keys),
            })
            total_savings = sum(float(r.get("annualSavings", r.get("savingsAmount", 0)) or 0) for r in cost_recs)
            desc = "Azure Advisor estimates potential savings."
            desc_es = "Azure Advisor estima posibles ahorros."
            if total_savings > 0:
                desc = f"Azure Advisor estimates ~${total_savings:,.0f} in potential annual savings."
                desc_es = f"Azure Advisor estima ~${total_savings:,.0f} en posibles ahorros anuales."
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "high",
                "title": f"{len(cost_recs)} Cost optimization recommendations",
                "title_es": f"{len(cost_recs)} Recomendaciones de optimización de costos",
                "description": desc,
                "description_es": desc_es,
                "resources": [r.get("impactedResource", r.get("impactedValue", "")) for r in cost_recs[:10]]
            })

        # Reservation / Savings Plan recommendations (rate optimization). Informational only -
        # these are opportunities on already-healthy resources, so they don't affect the score,
        # same treatment as the community checklist findings.
        reservation_recs = [
            r for r in self._data_rows(self.data.get("advisor", {}).get("ReservationRecommendations", []))
            if r.get("ResourceType") or r.get("SkuName")
        ]
        if reservation_recs:
            total_net_savings = sum(float(r.get("NetSavings", 0) or 0) for r in reservation_recs)
            currency = next((r.get("Currency") for r in reservation_recs if r.get("Currency")), "USD")
            desc = f"Buying the recommended Reserved Instances/Savings Plans could save ~{total_net_savings:,.0f} {currency}/year vs. pay-as-you-go."
            desc_es = f"Comprar las Reservas/Planes de Ahorro recomendados podría ahorrar ~{total_net_savings:,.0f} {currency}/año frente a pago por uso."
            self.action_items.append({
                "pillar": "Cost Optimization",
                "severity": "medium",
                "title": f"{len(reservation_recs)} Reservation/Savings Plan opportunities",
                "title_es": f"{len(reservation_recs)} oportunidades de Reservas/Planes de Ahorro",
                "description": desc,
                "description_es": desc_es,
                "resources": [f"{r.get('ResourceType', '')} {r.get('SkuName', '')} x{r.get('RecommendedQty', '?')} ({r.get('Term', '')})" for r in reservation_recs[:10]]
            })
        
        if total_pool > 0:
            impacted = min(len(impacted_keys), total_pool)
            healthy = max(0, total_pool - impacted)
            score = round((healthy / total_pool) * 100)
        else:
            healthy, score = 0, None
        
        self.scores["Cost Optimization"] = score if score is not None else 0
        self.breakdowns["Cost Optimization"] = {
            "available": score is not None,
            "sources": sources,
            "impacted": min(len(impacted_keys), total_pool),
            "healthy": healthy,
            "total": total_pool,
        }
    
    def _analyze_operational_excellence(self):
        """Operational Excellence score using the official Advisor Score subcategory model."""
        total_pool = self._resource_pool()
        synthetic = []
        
        # Diagnostic settings coverage
        diag = self.data.get("metrics", {}).get("DiagnosticsCoverage", [])
        if diag:
            no_diag = [d for d in diag if str(d.get("HasDiagnostics", d.get("Gap", ""))).lower() in ("false", "no diagnostics configured")]
            if no_diag:
                pct_missing = len(no_diag) / len(diag) * 100
                synthetic += [{"problem": "Resource missing diagnostic settings",
                               "impactedResource": d.get("Name", d.get("name", ""))} for d in no_diag]
                self.action_items.append({
                    "pillar": "Operational Excellence",
                    "severity": "high",
                    "title": f"{len(no_diag)}/{len(diag)} critical resources missing diagnostic settings ({pct_missing:.0f}%)",
                    "title_es": f"{len(no_diag)}/{len(diag)} recursos críticos sin configuración de diagnóstico ({pct_missing:.0f}%)",
                    "description": "Configure diagnostic settings to send logs to Log Analytics for observability and troubleshooting.",
                    "description_es": "Configure ajustes de diagnóstico para enviar registros a Log Analytics para observabilidad y solución de problemas.",
                    "resources": [d.get("Name", d.get("name", "")) for d in no_diag[:10]]
                })
        
        # Tag coverage — computed from the per-resource `tags` field across Discovery inventory
        # sheets (there's no dedicated tag-usage export; GovViz's Subscriptions.tags is subscription-level only).
        TAGGABLE_SHEETS = ["VMs", "AppServices", "AKS", "VNets", "NSGs", "LoadBalancers",
                           "Firewalls", "PublicIPs", "Storage", "Databases", "KeyVaults"]
        discovery = self.data.get("discovery", {})
        total_taggable, untagged_names = 0, []
        for sheet in TAGGABLE_SHEETS:
            for r in discovery.get(sheet, []):
                total_taggable += 1
                tags = r.get("tags")
                if not tags or str(tags).strip().lower() in ("", "none", "{}"):
                    untagged_names.append(r.get("name", r.get("Name", "")))
        if total_taggable > 0 and untagged_names:
            pct_untagged = len(untagged_names) / total_taggable * 100
            synthetic += [{"problem": "Resource has no tags configured", "impactedResource": n} for n in untagged_names]
            self.action_items.append({
                "pillar": "Operational Excellence",
                "severity": "medium",
                "title": f"{len(untagged_names)}/{total_taggable} resources have no tags ({pct_untagged:.0f}%)",
                "title_es": f"{len(untagged_names)}/{total_taggable} recursos no tienen etiquetas ({pct_untagged:.0f}%)",
                "description": "Implement a tagging strategy (Owner, CostCenter, Environment, Application) for governance and cost allocation.",
                "description_es": "Implemente una estrategia de etiquetado (Owner, CostCenter, Environment, Application) para gobernanza y asignación de costos.",
                "resources": untagged_names[:10]
            })
        
        # Resource locks (protects against accidental delete/modify of critical resources)
        locks = self.data.get("governance", {}).get("Locks", [])
        has_locks = locks and not (len(locks) == 1 and locks[0].get("Result"))
        if not has_locks:
            self.action_items.append({
                "pillar": "Operational Excellence",
                "severity": "medium",
                "title": "No resource locks configured",
                "title_es": "No hay bloqueos de recursos configurados",
                "description": "Apply CanNotDelete or ReadOnly locks to critical resources (networking, databases, key vaults) to prevent accidental deletion or modification.",
                "description_es": "Aplique bloqueos CanNotDelete o ReadOnly a los recursos críticos (redes, bases de datos, key vaults) para evitar eliminaciones o modificaciones accidentales.",
                "resources": []
            })
        
        # Policy compliance
        compliance = self.data.get("governance", {}).get("PolicyCompliance", [])
        if compliance:
            total_nc = sum(int(c.get("NonCompliantCount", 0) or 0) for c in compliance)
            non_compliant = [c for c in compliance if int(c.get("NonCompliantCount", 0) or 0) > 0]
            synthetic += [{"problem": "Non-compliant policy evaluation",
                           "impactedResource": c.get("policyAssignment", c.get("policyAssignmentName", ""))} for c in non_compliant]
            if total_nc > 0:
                self.action_items.append({
                    "pillar": "Operational Excellence",
                    "severity": "medium",
                    "title": f"{total_nc} non-compliant policy evaluations",
                    "title_es": f"{total_nc} evaluaciones de políticas no conformes",
                    "description": "Review policy compliance and remediate non-compliant resources or adjust policy assignments.",
                    "description_es": "Revise el cumplimiento de políticas y corrija los recursos no conformes o ajuste las asignaciones de políticas.",
                    "resources": [f"{c.get('policyAssignment', c.get('policyAssignmentName', ''))} ({c.get('NonCompliantCount', 0)} violations)" for c in compliance[:10]]
                })
        
        # OpEx Advisor recommendations
        opex_recs = self.data.get("advisor", {}).get("OperationalExcellence", [])
        if opex_recs:
            self.action_items.append({
                "pillar": "Operational Excellence",
                "severity": "medium",
                "title": f"{len(opex_recs)} Operational Excellence recommendations from Advisor",
                "title_es": f"{len(opex_recs)} Recomendaciones de Excelencia Operativa de Advisor",
                "description": "Azure Advisor has identified operational improvements.",
                "description_es": "Azure Advisor identificó mejoras operativas.",
                "resources": [r.get("impactedResource", r.get("impactedValue", "")) for r in opex_recs[:10]]
            })

        buckets = AdvisorScoreModel.merge_buckets(
            AdvisorScoreModel.bucket(opex_recs, AdvisorScoreModel.OPERATIONAL_RULES),
            AdvisorScoreModel.bucket(synthetic, AdvisorScoreModel.OPERATIONAL_RULES),
        )
        score, rows = AdvisorScoreModel.score_category(AdvisorScoreModel.OPERATIONAL_WEIGHTS, buckets, total_pool)
        self.scores["Operational Excellence"] = score if score is not None else 0
        self.breakdowns["Operational Excellence"] = {"available": score is not None, "rows": rows, "pool": total_pool}
    
    def _analyze_performance(self):
        """Performance score using the official Advisor Score subcategory model."""
        total_pool = self._resource_pool()
        synthetic = []
        
        # Saturated VMs
        vm_metrics = self.data.get("metrics", {}).get("VM_RightSizing", [])
        saturated = [v for v in vm_metrics if "Saturated" in str(v.get("Assessment", ""))]
        synthetic += [{"problem": "Virtual machine CPU saturated",
                       "impactedResource": v.get("Name", "")} for v in saturated]
        if saturated:
            self.action_items.append({
                "pillar": "Performance Efficiency",
                "severity": "high",
                "title": f"{len(saturated)} VMs are saturated (>80% CPU)",
                "title_es": f"{len(saturated)} VMs están saturadas (>80% CPU)",
                "description": "These VMs are consistently at high CPU usage. Consider scaling up or scaling out.",
                "description_es": "Estas VMs presentan un uso de CPU consistentemente alto. Considere escalar verticalmente u horizontalmente.",
                "resources": [f"{v.get('Name', '')} ({v.get('VMSize', '')} @ {v.get('AvgCPU_Pct', '?')}% CPU)" for v in saturated[:10]]
            })
        
        # Saturated SQL
        sql_metrics = self.data.get("metrics", {}).get("SQL_RightSizing", [])
        saturated_sql = [s for s in sql_metrics if "Saturated" in str(s.get("Assessment", ""))]
        synthetic += [{"problem": "SQL database DTU/CPU saturated",
                       "impactedResource": s.get("Name", "")} for s in saturated_sql]
        if saturated_sql:
            self.action_items.append({
                "pillar": "Performance Efficiency",
                "severity": "high",
                "title": f"{len(saturated_sql)} SQL Databases are saturated (>80% DTU/CPU)",
                "title_es": f"{len(saturated_sql)} Bases de datos SQL están saturadas (>80% DTU/CPU)",
                "description": "These databases are consistently at high utilization. Consider scaling up the tier.",
                "description_es": "Estas bases de datos presentan una utilización consistentemente alta. Considere aumentar el nivel de servicio.",
                "resources": [f"{s.get('Name', '')} ({s.get('SKU', '')} @ {s.get('AvgUsage_Pct', '?')}%)" for s in saturated_sql[:10]]
            })
        
        # Performance advisor recommendations
        perf_recs = self.data.get("advisor", {}).get("Performance", [])
        if perf_recs:
            self.action_items.append({
                "pillar": "Performance Efficiency",
                "severity": "medium",
                "title": f"{len(perf_recs)} Performance recommendations from Advisor",
                "title_es": f"{len(perf_recs)} Recomendaciones de Rendimiento de Advisor",
                "description": "Azure Advisor has identified performance improvements.",
                "description_es": "Azure Advisor identificó mejoras de rendimiento.",
                "resources": [r.get("impactedResource", r.get("impactedValue", "")) for r in perf_recs[:10]]
            })

        buckets = AdvisorScoreModel.merge_buckets(
            AdvisorScoreModel.bucket(perf_recs, AdvisorScoreModel.PERFORMANCE_RULES),
            AdvisorScoreModel.bucket(synthetic, AdvisorScoreModel.PERFORMANCE_RULES),
        )
        score, rows = AdvisorScoreModel.score_category(AdvisorScoreModel.PERFORMANCE_WEIGHTS, buckets, total_pool)
        self.scores["Performance Efficiency"] = score if score is not None else 0
        self.breakdowns["Performance Efficiency"] = {"available": score is not None, "rows": rows, "pool": total_pool}
    
    def get_overall_score(self) -> int:
        """Mean of the available category scores (per the official Advisor Score model)."""
        available = [p for p, b in self.breakdowns.items() if b.get("available")]
        if not available:
            return 0
        return round(sum(self.scores[p] for p in available) / len(available))


# ═══════════════════════════════════════════════════════════════════════════════
# HTML DASHBOARD GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

class DashboardGenerator:
    """Generates the final consolidated HTML dashboard."""
    
    def __init__(self, engine: InsightEngine, data: dict, base_dir: Optional[str] = None):
        self.engine = engine
        self.data = data
        self.base_dir = base_dir

    @staticmethod
    def _count_real(rows: list) -> int:
        """Count sheet rows, ignoring the Export-Sheet 'No data found' placeholder row."""
        return len([
            row for row in rows
            if not row.get("Result") and str(row.get("DataStatus", "")).lower() != "nodata"
        ])

    _ICON_FILE_CACHE: dict = {}
    _icon_instance_counter = 0

    def _azure_icon(self, name: str, css_class: str = "az-icon") -> str:
        """Inline an official Microsoft Azure architecture SVG icon (assets/azure-icons/*.svg).

        Icon ids/gradient refs are namespaced per-embed since the same icon may be
        rendered many times in one page (duplicate SVG ids are otherwise invalid HTML).
        """
        svg = self._ICON_FILE_CACHE.get(name)
        if svg is None:
            icon_path = Path(__file__).parent / "assets" / "azure-icons" / f"{name}.svg"
            svg = icon_path.read_text(encoding="utf-8")
            self._ICON_FILE_CACHE[name] = svg
        DashboardGenerator._icon_instance_counter += 1
        uid = DashboardGenerator._icon_instance_counter
        svg = re.sub(r'id="([^"]+)"', lambda m: f'id="{m.group(1)}-{uid}"', svg)
        svg = re.sub(r'url\(#([^)]+)\)', lambda m: f'url(#{m.group(1)}-{uid})', svg)
        svg = re.sub(r"<svg ", f'<svg class="{css_class}" ', svg, count=1)
        return svg

    def _render_security_view(self) -> str:
        """Render dedicated security posture, compliance, and operations details."""
        security = self.data.get("security", {})
        breakdown = self.engine.breakdowns.get("Security", {})

        def real_rows(sheet_name):
            return [
                row for row in security.get(sheet_name, [])
                if not row.get("Result") and str(row.get("DataStatus", "")).lower() != "nodata"
            ]

        def as_int(value):
            try:
                return int(value or 0)
            except (TypeError, ValueError):
                return 0

        def render_table(rows, columns, empty_en, empty_es, limit=12):
            if not rows:
                return f'<p class="empty-state">{bi(empty_en, empty_es)}</p>'
            header = "".join(f"<th>{bi(label_en, label_es)}</th>" for _, label_en, label_es, _ in columns)
            body = ""
            for i, row in enumerate(rows):
                cells = "".join(
                    f'<td class="{"num" if numeric else ""}">{escape(str(row.get(key, "")))}</td>'
                    for key, _, _, numeric in columns
                )
                row_attr = ' class="extra-row" hidden' if i >= limit else ""
                body += f"<tr{row_attr}>{cells}</tr>"
            more = ""
            if len(rows) > limit:
                more_text = bi(f"+ {len(rows) - limit} more", f"+ {len(rows) - limit} m\u00e1s")
                less_text = bi("Show less", "Mostrar menos")
                more = (
                    '<button type="button" class="table-more-note" onclick="toggleTableRows(this)">'
                    f'<span class="more-label">{more_text}</span><span class="less-label">{less_text}</span></button>'
                )
            return f'<div class="security-table-wrap"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>{more}'

        severity_rank = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
        recommendations = [
            row for row in real_rows("Recommendations")
            if str(row.get("recommendationState", "")).lower() not in ("healthy", "notapplicable")
        ]
        recommendations.sort(key=lambda row: severity_rank.get(str(row.get("recommendationSeverity", "")).lower(), 5))
        controls = sorted(
            real_rows("ScoreControls"),
            key=lambda row: as_int(row.get("unhealthyResourceCount")),
            reverse=True,
        )
        mcsb = sorted(
            real_rows("MCSBCompliance"),
            key=lambda row: as_int(row.get("failedResources")),
            reverse=True,
        )
        defender_plans = real_rows("DefenderPlans")
        source_status = real_rows("SourceStatus")
        incidents = sorted(real_rows("Incidents"), key=lambda row: str(row.get("LastUpdateDateTime", "")), reverse=True)
        alerts = sorted(real_rows("Alerts"), key=lambda row: str(row.get("LastUpdateDateTime", "")), reverse=True)
        endpoint_recommendations = sorted(
            real_rows("EndpointRecommendations"),
            key=lambda row: as_int(row.get("ExposedMachinesCount")),
            reverse=True,
        )
        vulnerabilities = sorted(
            real_rows("Vulnerabilities"),
            key=lambda row: severity_rank.get(str(row.get("Severity", "")).lower(), 5),
        )
        machines = real_rows("Machines")
        at_risk_machines = len([
            machine for machine in machines
            if str(machine.get("RiskScore", "")).lower() in ("critical", "high")
        ])

        secure_score = breakdown.get("secure_score_pct")
        secure_score_text = f"{secure_score:.1f}%" if secure_score is not None else "N/A"
        stats = [
            (secure_score_text, bi("Defender Secure Score", "Secure Score de Defender"), "security-accent"),
            (breakdown.get("open_recommendations", 0), bi("Open CSPM Recommendations", "Recomendaciones CSPM abiertas"), "security-danger"),
            (breakdown.get("mcsb_failed_resources", 0), bi("Failed MCSB Resources", "Recursos con error en MCSB"), "security-warning"),
            (breakdown.get("active_incidents", 0), bi("Active XDR Incidents", "Incidentes XDR activos"), "security-danger"),
            (breakdown.get("active_alerts", 0), bi("Active Security Alerts", "Alertas de seguridad activas"), "security-warning"),
            (at_risk_machines, bi("High-risk Endpoints", "Endpoints de alto riesgo"), "security-danger"),
        ]
        stat_html = "".join(
            f'<div class="security-stat"><div class="value {color_class}">{value}</div><div class="label">{label}</div></div>'
            for value, label, color_class in stats
        )

        status_rows = ""
        for row in source_status:
            status = str(row.get("Status", "Unknown"))
            status_class = re.sub(r"[^a-z]", "", status.lower())
            status_rows += f"""<tr>
                <td>{escape(str(row.get('Source', '')))}</td>
                <td><span class="source-status source-{status_class}">{escape(status)}</span></td>
                <td class="num">{escape(str(row.get('Records', 0)))}</td>
                <td>{escape(str(row.get('Details', '')))}</td>
            </tr>"""
        if source_status:
            source_table = f"""<div class="security-table-wrap"><table>
                <thead><tr><th>{bi('Source', 'Fuente')}</th><th>{bi('Status', 'Estado')}</th><th>{bi('Records', 'Registros')}</th><th>{bi('Details', 'Detalles')}</th></tr></thead>
                <tbody>{status_rows}</tbody>
            </table></div>"""
        else:
            source_table = f'<p class="empty-state">{bi("Security source status is unavailable.", "El estado de las fuentes de seguridad no está disponible.")}</p>'

        recommendations_table = render_table(recommendations, [
            ("recommendationSeverity", "Severity", "Severidad", False),
            ("recommendationName", "Recommendation", "Recomendación", False),
            ("recommendationState", "State", "Estado", False),
            ("affectedResourceId", "Affected Resource", "Recurso afectado", False),
        ], "No open Defender for Cloud recommendations.", "No hay recomendaciones abiertas de Defender for Cloud.")
        controls_table = render_table(controls, [
            ("controlName", "Secure Score Control", "Control de Secure Score", False),
            ("healthyResourceCount", "Healthy", "Saludables", True),
            ("unhealthyResourceCount", "Unhealthy", "No saludables", True),
            ("currentScore", "Current Score", "Puntaje actual", True),
            ("maxScore", "Max Score", "Puntaje máximo", True),
        ], "No Secure Score controls returned.", "No se devolvieron controles de Secure Score.")
        mcsb_table = render_table(mcsb, [
            ("complianceControl", "MCSB Control", "Control MCSB", False),
            ("assessmentName", "Assessment", "Evaluación", False),
            ("state", "State", "Estado", False),
            ("passedResources", "Passed", "Aprobados", True),
            ("failedResources", "Failed", "Con error", True),
        ], "No MCSB assessment data returned.", "No se devolvieron datos de evaluación MCSB.")
        plans_table = render_table(defender_plans, [
            ("planName", "Defender Plan", "Plan de Defender", False),
            ("pricingTier", "Tier", "Nivel", False),
            ("subPlan", "Sub-plan", "Subplan", False),
            ("subscriptionId", "Subscription", "Suscripción", False),
        ], "No Defender plan data returned.", "No se devolvieron datos de planes de Defender.")
        incidents_table = render_table(incidents, [
            ("Severity", "Severity", "Severidad", False),
            ("DisplayName", "Incident", "Incidente", False),
            ("Status", "Status", "Estado", False),
            ("AssignedTo", "Assigned To", "Asignado a", False),
            ("LastUpdateDateTime", "Last Updated", "Última actualización", False),
        ], "No incidents returned for the selected lookback period.", "No se devolvieron incidentes para el período seleccionado.")
        alerts_table = render_table(alerts, [
            ("Severity", "Severity", "Severidad", False),
            ("Title", "Alert", "Alerta", False),
            ("Status", "Status", "Estado", False),
            ("ServiceSource", "Service", "Servicio", False),
            ("LastUpdateDateTime", "Last Updated", "Última actualización", False),
        ], "No alerts returned for the selected lookback period.", "No se devolvieron alertas para el período seleccionado.")
        endpoint_recommendations_table = render_table(endpoint_recommendations, [
            ("RecommendationName", "Endpoint Recommendation", "Recomendación de Endpoint", False),
            ("ExposedMachinesCount", "Exposed Machines", "Equipos expuestos", True),
            ("SeverityScore", "Severity Score", "Puntaje de severidad", True),
            ("Status", "Status", "Estado", False),
        ], "No Defender for Endpoint recommendations returned.", "No se devolvieron recomendaciones de Defender for Endpoint.")
        vulnerabilities_table = render_table(vulnerabilities, [
            ("VulnerabilityId", "CVE", "CVE", False),
            ("Severity", "Severity", "Severidad", False),
            ("CvssV3", "CVSS v3", "CVSS v3", True),
            ("ExposedMachines", "Exposed Machines", "Equipos expuestos", True),
            ("PublicExploit", "Public Exploit", "Exploit público", False),
        ], "No Defender for Endpoint vulnerabilities returned.", "No se devolvieron vulnerabilidades de Defender for Endpoint.")

        return f"""
        <div class="security-summary-band">
            <div>
                <div class="security-eyebrow">MICROSOFT DEFENDER</div>
                <h2>{bi('Security posture and operations', 'Postura y operaciones de seguridad')}</h2>
                <p>{bi('CSPM, Microsoft Cloud Security Benchmark, Defender XDR, and endpoint exposure in one assessment.', 'CSPM, Microsoft Cloud Security Benchmark, Defender XDR y exposición de endpoints en una sola evaluación.')}</p>
            </div>
            <div class="security-score-mark">{secure_score_text}</div>
        </div>
        <div class="security-stats">{stat_html}</div>

        <div class="section">
            <div class="section-title security-title">{bi('CSPM recommendations', 'Recomendaciones CSPM')}</div>
            {recommendations_table}
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Secure Score controls', 'Controles de Secure Score')}</h3>{controls_table}</div>
            <div class="security-panel"><h3>{bi('Microsoft Cloud Security Benchmark', 'Microsoft Cloud Security Benchmark')}</h3>{mcsb_table}</div>
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Defender for Cloud coverage', 'Cobertura de Defender for Cloud')}</h3>{plans_table}</div>
            <div class="security-panel"><h3>{bi('Data source coverage', 'Cobertura de fuentes de datos')}</h3>{source_table}</div>
        </div>
        <div class="section">
            <div class="section-title security-title">{bi('Defender XDR operations', 'Operaciones de Defender XDR')}</div>
            <div class="security-two-column">
                <div class="security-panel"><h3>{bi('Recent incidents', 'Incidentes recientes')}</h3>{incidents_table}</div>
                <div class="security-panel"><h3>{bi('Recent alerts', 'Alertas recientes')}</h3>{alerts_table}</div>
            </div>
        </div>
        <div class="section">
            <div class="section-title security-title">{bi('Defender for Endpoint exposure', 'Exposición de Defender for Endpoint')}</div>
            <div class="security-two-column">
                <div class="security-panel"><h3>{bi('Security recommendations', 'Recomendaciones de seguridad')}</h3>{endpoint_recommendations_table}</div>
                <div class="security-panel"><h3>{bi('Vulnerabilities', 'Vulnerabilidades')}</h3>{vulnerabilities_table}</div>
            </div>
        </div>"""

    def _render_governance_view(self) -> str:
        """Render management-group hierarchy, policy/RBAC, and scope details natively (no iframe to a separate HTML file)."""
        gov = self.data.get("governance", {})

        def real_rows(sheet_name):
            return [
                row for row in gov.get(sheet_name, [])
                if not row.get("Result") and str(row.get("DataStatus", "")).lower() != "nodata"
            ]

        def render_table(rows, columns, empty_en, empty_es, limit=15):
            if not rows:
                return f'<p class="empty-state">{bi(empty_en, empty_es)}</p>'
            header = "".join(f"<th>{bi(label_en, label_es)}</th>" for _, label_en, label_es, _ in columns)
            body = ""
            for i, row in enumerate(rows):
                cells = "".join(
                    f'<td class="{"num" if numeric else ""}">{escape(str(row.get(key, "")))}</td>'
                    for key, _, _, numeric in columns
                )
                row_attr = ' class="extra-row" hidden' if i >= limit else ""
                body += f"<tr{row_attr}>{cells}</tr>"
            more = ""
            if len(rows) > limit:
                more_text = bi(f"+ {len(rows) - limit} more", f"+ {len(rows) - limit} m\u00e1s")
                less_text = bi("Show less", "Mostrar menos")
                more = (
                    '<button type="button" class="table-more-note" onclick="toggleTableRows(this)">'
                    f'<span class="more-label">{more_text}</span><span class="less-label">{less_text}</span></button>'
                )
            return f'<div class="security-table-wrap"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>{more}'

        mg_rows = real_rows("MgmtGroups")
        subs_rows = real_rows("Subscriptions")
        policy_assignments = real_rows("PolicyAssignments")
        custom_policies = real_rows("CustomPolicies")
        policy_compliance = real_rows("PolicyCompliance")
        role_assignments = real_rows("RoleAssignments")
        custom_roles = real_rows("CustomRoles")
        defender_plans = real_rows("DefenderPlans")
        secure_scores = real_rows("SecureScores")
        resource_summary = real_rows("ResourceSummary")
        orphaned = real_rows("OrphanedResources")
        vnets = real_rows("VNets")
        locks = real_rows("Locks")

        non_compliant_count = sum(int(r.get("NonCompliantCount", 0) or 0) for r in policy_compliance) or len(policy_compliance)
        stats = [
            (len(mg_rows), bi("Management Groups", "Grupos de Administraci\u00f3n")),
            (len(subs_rows), bi("Subscriptions", "Suscripciones")),
            (len(policy_assignments), bi("Policy Assignments", "Asignaciones de Pol\u00edticas")),
            (non_compliant_count, bi("Non-Compliant Resources", "Recursos No Conformes")),
            (len(role_assignments), bi("Role Assignments", "Asignaciones de Roles")),
            (len(locks), bi("Resource Locks", "Bloqueos de Recursos")),
            (len(orphaned), bi("Orphaned Resources", "Recursos Hu\u00e9rfanos")),
        ]
        stat_html = "".join(
            f'<div class="security-stat"><div class="value security-accent">{count:,}</div><div class="label">{label}</div></div>'
            for count, label in stats
        )

        def scope_matches(scope, mg_id):
            return bool(mg_id) and f"/managementgroups/{mg_id}".lower() in str(scope or "").lower()

        def mg_level(m):
            try:
                return int(m.get("Level"))
            except (TypeError, ValueError):
                return None

        has_hierarchy = bool(mg_rows) and any(mg_level(m) is not None for m in mg_rows)

        def build_tree(mg, depth_guard=0):
            if depth_guard > 12:
                return ""
            mg_id = str(mg.get("Id", ""))
            level = mg_level(mg)
            level = level if level is not None else 0
            path = str(mg.get("Path", "") or "")
            policy_count = len([p for p in policy_assignments if scope_matches(p.get("scope_"), mg_id)])
            role_count = len([r for r in role_assignments if scope_matches(r.get("scope_"), mg_id)])
            sub_count = int(mg.get("Subscriptions", 0) or 0)
            children = [m for m in mg_rows if mg_level(m) == level + 1 and str(m.get("Path", "")).startswith(f"{path}/")]
            children_html = "".join(build_tree(child, depth_guard + 1) for child in children)
            sub_html = ""
            if sub_count:
                sub_names = escape(str(mg.get("SubNames", "") or ""))
                sub_html = f'<li class="mg-sub-leaf" title="{sub_names}">{self._azure_icon("subscriptions")} {sub_count} {bi("subscriptions", "suscripciones")}</li>'
            return f"""<li>
                <details open class="mg-node-details">
                    <summary class="mg-node">
                        <span class="mg-name">{self._azure_icon("management-groups")} {escape(str(mg.get('DisplayName', mg_id)))}</span>
                        <span class="mg-badge mg-badge-policy" title="{policy_count} policy assignments">{self._azure_icon("policy")} {policy_count}</span>
                        <span class="mg-badge mg-badge-rbac" title="{role_count} role assignments">{role_count}</span>
                    </summary>
                    <ul>{children_html}{sub_html}</ul>
                </details>
            </li>"""

        if has_hierarchy:
            roots = [m for m in mg_rows if mg_level(m) == 0]
            hierarchy_html = f'<ul class="mg-tree">{"".join(build_tree(r) for r in roots)}</ul>' if roots else ""
        else:
            hierarchy_html = ""
        if not hierarchy_html:
            hierarchy_html = render_table(mg_rows, [
                ("displayName", "Name", "Nombre", False),
                ("name", "Id", "Id", False),
                ("parent", "Parent", "Padre", False),
            ], "No management group data found.", "No se encontraron datos de grupos de administraci\u00f3n.")

        scope_items = []
        if has_hierarchy:
            ordered_mgs = sorted(mg_rows, key=lambda m: (mg_level(m) or 0, str(m.get("DisplayName", ""))))
            for mg in ordered_mgs[:40]:
                mg_id = str(mg.get("Id", ""))
                mg_policies = [p for p in policy_assignments if scope_matches(p.get("scope_"), mg_id)]
                mg_roles = [r for r in role_assignments if scope_matches(r.get("scope_"), mg_id)]
                indent = "\u2014" * (mg_level(mg) or 0)
                if mg_policies:
                    policy_rows = "".join(
                        f"<tr><td>{escape(str(p.get('displayName', '')))}</td><td>{escape(str(p.get('enforcement', '')))}</td><td>{escape(str(p.get('identity_', '')))}</td></tr>"
                        for p in mg_policies[:10]
                    )
                    policy_table = f"""<table class="mini-table"><thead><tr><th>{bi('Policy', 'Pol\u00edtica')}</th><th>{bi('Enforcement', 'Aplicaci\u00f3n')}</th><th>{bi('Identity', 'Identidad')}</th></tr></thead><tbody>{policy_rows}</tbody></table>"""
                else:
                    policy_table = f'<p class="empty-state">{bi("No policy assignments at this scope.", "Sin asignaciones de pol\u00edticas en este alcance.")}</p>'
                scope_items.append(f"""<details class="scope-detail">
                    <summary>{indent} <strong>{escape(str(mg.get('DisplayName', mg_id)))}</strong> <span class="scope-id">({escape(mg_id)})</span>
                        <span class="mg-badge mg-badge-policy">{len(mg_policies)} {bi('policies', 'pol.')}</span>
                        <span class="mg-badge mg-badge-rbac">{len(mg_roles)} {bi('roles', 'roles')}</span>
                        <span class="mg-badge mg-badge-subs">{mg.get('Subscriptions', 0)} {bi('subs', 'subs')}</span>
                    </summary>
                    <p><strong>{bi('Path', 'Ruta')}:</strong> {escape(str(mg.get('Path', '')))}</p>
                    {policy_table}
                </details>""")
        scope_insights_html = "".join(scope_items) if scope_items else f'<p class="empty-state">{bi("Scope insights require the Management Group hierarchy (Level/Path fields).", "Los detalles de alcance requieren la jerarqu\u00eda de Grupos de Administraci\u00f3n (campos Level/Path).")}</p>'

        policy_assignments_table = render_table(policy_assignments, [
            ("displayName", "Name", "Nombre", False),
            ("enforcement", "Enforcement", "Aplicaci\u00f3n", False),
            ("identity_", "Identity", "Identidad", False),
            ("scope_", "Scope", "Alcance", False),
        ], "No policy assignments found.", "No se encontraron asignaciones de pol\u00edticas.")
        custom_policies_table = render_table(custom_policies, [
            ("displayName", "Name", "Nombre", False),
            ("effect", "Effect", "Efecto", False),
            ("category", "Category", "Categor\u00eda", False),
        ], "No custom policy definitions found.", "No se encontraron definiciones de pol\u00edticas personalizadas.")
        policy_compliance_table = render_table(policy_compliance, [
            ("policyAssignment", "Policy Assignment", "Asignaci\u00f3n de Pol\u00edtica", False),
            ("complianceState", "State", "Estado", False),
            ("NonCompliantCount", "Non-Compliant", "No Conformes", True),
        ], "No non-compliant resources found.", "No se encontraron recursos no conformes.")
        role_assignments_table = render_table(role_assignments, [
            ("principalType", "Principal Type", "Tipo de Principal", False),
            ("roleDefId", "Role", "Rol", False),
            ("scope_", "Scope", "Alcance", False),
        ], "No role assignments found.", "No se encontraron asignaciones de roles.")
        custom_roles_table = render_table(custom_roles, [
            ("roleName", "Role Name", "Nombre del Rol", False),
            ("description_", "Description", "Descripci\u00f3n", False),
        ], "No custom role definitions found.", "No se encontraron definiciones de roles personalizados.")
        defender_plans_table = render_table(defender_plans, [
            ("name", "Plan", "Plan", False),
            ("tier", "Tier", "Nivel", False),
            ("subPlan", "Sub-plan", "Sub-plan", False),
        ], "No Defender for Cloud plans found.", "No se encontraron planes de Defender for Cloud.")
        secure_scores_table = render_table(secure_scores, [
            ("subscriptionId", "Subscription", "Suscripci\u00f3n", False),
            ("pct", "Score %", "Puntaje %", True),
        ], "No Secure Score data found.", "No se encontraron datos de Secure Score.")
        resource_summary_table = render_table(resource_summary, [
            ("type", "Resource Type", "Tipo de Recurso", False),
            ("location", "Location", "Ubicaci\u00f3n", False),
            ("Count", "Count", "Cantidad", True),
        ], "No resource summary found.", "No se encontr\u00f3 resumen de recursos.")
        orphaned_table = render_table(orphaned, [
            ("Type", "Type", "Tipo", False),
            ("Name", "Name", "Nombre", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("Detail", "Detail", "Detalle", False),
        ], "No orphaned resources found.", "No se encontraron recursos hu\u00e9rfanos.")
        vnets_table = render_table(vnets, [
            ("name", "VNet", "VNet", False),
            ("addressSpace", "Address Space", "Espacio de Direcciones", False),
            ("subnets", "Subnets", "Subredes", True),
            ("peerings", "Peerings", "Emparejamientos", True),
        ], "No virtual networks found.", "No se encontraron redes virtuales.")
        locks_table = render_table(locks, [
            ("name", "Name", "Nombre", False),
            ("lockLevel", "Level", "Nivel", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
        ], "No resource locks found.", "No se encontraron bloqueos de recursos.")

        return f"""
        <div class="security-summary-band">
            <div>
                <div class="security-eyebrow">GOVERNANCE VISUALIZER</div>
                <h2>{bi('Management hierarchy, policy, and RBAC', 'Jerarqu\u00eda de administraci\u00f3n, pol\u00edticas y RBAC')}</h2>
                <p>{bi('Management group tree, policy/RBAC assignments, and per-scope insights in one native view.', 'Árbol de grupos de administraci\u00f3n, asignaciones de pol\u00edticas/RBAC y detalles por alcance en una vista nativa.')}</p>
            </div>
        </div>
        <div class="security-stats">{stat_html}</div>

        <div class="section">
            <div class="section-title security-title">{bi('Hierarchy map', 'Mapa de jerarqu\u00eda')}</div>
            {hierarchy_html}
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Policy assignments', 'Asignaciones de pol\u00edticas')}</h3>{policy_assignments_table}</div>
            <div class="security-panel"><h3>{bi('Custom policy definitions', 'Definiciones de pol\u00edticas personalizadas')}</h3>{custom_policies_table}</div>
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Non-compliant resources', 'Recursos no conformes')}</h3>{policy_compliance_table}</div>
            <div class="security-panel"><h3>{bi('Resource locks', 'Bloqueos de recursos')}</h3>{locks_table}</div>
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Role assignments', 'Asignaciones de roles')}</h3>{role_assignments_table}</div>
            <div class="security-panel"><h3>{bi('Custom role definitions', 'Definiciones de roles personalizados')}</h3>{custom_roles_table}</div>
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Defender for Cloud plans', 'Planes de Defender for Cloud')}</h3>{defender_plans_table}</div>
            <div class="security-panel"><h3>{bi('Secure Score by subscription', 'Secure Score por suscripci\u00f3n')}</h3>{secure_scores_table}</div>
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Resources by type & location', 'Recursos por tipo y ubicaci\u00f3n')}</h3>{resource_summary_table}</div>
            <div class="security-panel"><h3>{bi('Orphaned resources (cost savings)', 'Recursos hu\u00e9rfanos (ahorro de costos)')}</h3>{orphaned_table}</div>
        </div>
        <div class="section">
            <div class="section-title security-title">{bi('Virtual networks', 'Redes virtuales')}</div>
            {vnets_table}
        </div>
        <div class="section">
            <div class="section-title security-title">{bi('Scope insights', 'Detalles por alcance')}</div>
            <div class="scope-insights">{scope_insights_html}</div>
        </div>"""

    def _render_cost_view(self) -> str:
        """Render cost optimization insights: right-sizing, idle/orphaned assets, Advisor and reservation savings."""
        advisor = self.data.get("advisor", {})
        metrics = self.data.get("metrics", {})
        governance = self.data.get("governance", {})
        discovery = self.data.get("discovery", {})

        def real_rows(source, sheet_name):
            return [
                row for row in source.get(sheet_name, [])
                if not row.get("Result") and str(row.get("DataStatus", "")).lower() != "nodata"
            ]

        def as_float(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return 0.0

        def render_table(rows, columns, empty_en, empty_es, limit=15):
            if not rows:
                return f'<p class="empty-state">{bi(empty_en, empty_es)}</p>'
            header = "".join(f"<th>{bi(label_en, label_es)}</th>" for _, label_en, label_es, _ in columns)
            body = ""
            for i, row in enumerate(rows):
                cells = "".join(
                    f'<td class="{"num" if numeric else ""}">{escape(str(row.get(key, "")))}</td>'
                    for key, _, _, numeric in columns
                )
                row_attr = ' class="extra-row" hidden' if i >= limit else ""
                body += f"<tr{row_attr}>{cells}</tr>"
            more = ""
            if len(rows) > limit:
                more_text = bi(f"+ {len(rows) - limit} more", f"+ {len(rows) - limit} m\u00e1s")
                less_text = bi("Show less", "Mostrar menos")
                more = (
                    '<button type="button" class="table-more-note" onclick="toggleTableRows(this)">'
                    f'<span class="more-label">{more_text}</span><span class="less-label">{less_text}</span></button>'
                )
            return f'<div class="security-table-wrap"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>{more}'

        def waste_priority(assessment):
            label = str(assessment or "").lower()
            if "idle" in label or "zero" in label:
                return 0
            if "underutil" in label or "oversiz" in label or "minimal" in label:
                return 1
            if "no data" in label:
                return 3
            return 2

        def assessment_color(assessment):
            label = str(assessment or "").lower()
            if "idle" in label or "zero" in label:
                return "var(--cp-danger)"
            if "underutil" in label or "oversiz" in label or "minimal" in label:
                return "var(--cp-warning)"
            if "no data" in label:
                return "var(--cp-text-muted)"
            return "var(--cp-success)"

        def render_distribution(rows):
            if not rows:
                return ""
            counts = {}
            for row in rows:
                label = str(row.get("Assessment", "Unknown")) or "Unknown"
                counts[label] = counts.get(label, 0) + 1
            total = len(rows)
            bars = "".join(
                f"""<div class="cost-dist-row">
                    <span class="cost-dist-label">{escape(label)}</span>
                    <div class="cost-dist-track"><div class="cost-dist-fill" style="width:{round(count / total * 100)}%; background:{assessment_color(label)};"></div></div>
                    <span class="cost-dist-count">{count}</span>
                </div>"""
                for label, count in sorted(counts.items(), key=lambda kv: kv[1], reverse=True)
            )
            return f'<div class="cost-dist">{bars}</div>'

        cost_recs = real_rows(advisor, "Cost")
        reservation_recs = [r for r in real_rows(advisor, "ReservationRecommendations") if r.get("ResourceType") or r.get("SkuName")]
        vm_rightsizing = sorted(real_rows(metrics, "VM_RightSizing"), key=lambda r: waste_priority(r.get("Assessment")))
        sql_rightsizing = sorted(real_rows(metrics, "SQL_RightSizing"), key=lambda r: waste_priority(r.get("Assessment")))
        plan_rightsizing = sorted(real_rows(metrics, "AppPlan_RightSizing"), key=lambda r: waste_priority(r.get("Assessment")))
        storage_activity = sorted(real_rows(metrics, "Storage_Activity"), key=lambda r: waste_priority(r.get("Assessment")))
        orphaned = real_rows(governance, "OrphanedResources")
        dealloc_vms = real_rows(discovery, "DeallocatedVMs")
        unattached_disks = real_rows(discovery, "UnattachedDisks")

        vm_waste = [r for r in vm_rightsizing if waste_priority(r.get("Assessment")) <= 1]
        sql_waste = [r for r in sql_rightsizing if waste_priority(r.get("Assessment")) <= 1]
        plan_waste = [r for r in plan_rightsizing if waste_priority(r.get("Assessment")) <= 1]
        storage_waste = [r for r in storage_activity if waste_priority(r.get("Assessment")) <= 1]

        advisor_savings = sum(as_float(r.get("annualSavings")) for r in cost_recs)
        reservation_savings = sum(as_float(r.get("NetSavings")) for r in reservation_recs)
        total_savings = advisor_savings + reservation_savings
        currency = next((r.get("Currency") for r in reservation_recs if r.get("Currency")), None) \
            or next((r.get("savingsCurrency") for r in cost_recs if r.get("savingsCurrency")), "USD")
        savings_text = f"${total_savings:,.0f}" if total_savings > 0 else "N/A"

        orphaned_total = len(orphaned) + len(dealloc_vms) + len(unattached_disks)

        stats = [
            (savings_text, bi("Est. Annual Savings", "Ahorro Anual Est."), "security-accent"),
            (len(cost_recs), bi("Advisor Cost Recs", "Recom. de Costos"), ""),
            (len(reservation_recs), bi("Reservation Opportunities", "Oportunidades de Reserva"), ""),
            (len(vm_waste), bi("Idle/Underutilized VMs", "VMs Inactivas/Subutilizadas"), "security-warning"),
            (len(sql_waste), bi("SQL DBs to Rightsize", "BD SQL para Ajustar"), "security-warning"),
            (len(plan_waste), bi("App Plans to Rightsize", "Planes App para Ajustar"), "security-warning"),
            (len(storage_waste), bi("Low-activity Storage", "Almacenamiento de Baja Actividad"), "security-warning"),
            (orphaned_total, bi("Orphaned/Unused Resources", "Recursos Hu\u00e9rfanos/No Usados"), "security-danger"),
        ]
        stat_html = "".join(
            f'<div class="security-stat"><div class="value {color_class}">{value}</div><div class="label">{label}</div></div>'
            for value, label, color_class in stats
        )

        for r in cost_recs:
            savings = as_float(r.get("annualSavings"))
            r["AnnualSavingsDisplay"] = f"{savings:,.0f} {r.get('savingsCurrency', '') or ''}".strip() if savings else ""
        for r in reservation_recs:
            savings = as_float(r.get("NetSavings"))
            r["NetSavingsDisplay"] = f"{savings:,.0f} {r.get('Currency', '') or ''}".strip() if savings else ""

        cost_recs_table = render_table(cost_recs, [
            ("impact", "Impact", "Impacto", False),
            ("problem", "Problem", "Problema", False),
            ("solution", "Solution", "Soluci\u00f3n", False),
            ("AnnualSavingsDisplay", "Annual Savings", "Ahorro Anual", True),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
        ], "No Advisor cost recommendations found.", "No se encontraron recomendaciones de costos de Advisor.")
        reservation_table = render_table(reservation_recs, [
            ("Subscription", "Subscription", "Suscripci\u00f3n", False),
            ("ResourceType", "Resource Type", "Tipo de Recurso", False),
            ("SkuName", "SKU", "SKU", False),
            ("Term", "Term", "Plazo", False),
            ("RecommendedQty", "Qty", "Cant.", True),
            ("NetSavingsDisplay", "Net Savings/yr", "Ahorro Neto/a\u00f1o", True),
        ], "No reservation/savings plan opportunities found.", "No se encontraron oportunidades de reservas/planes de ahorro.")
        vm_table = render_table(vm_rightsizing, [
            ("Name", "VM", "VM", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("VMSize", "Size", "Tama\u00f1o", False),
            ("AvgCPU_Pct", "Avg CPU %", "CPU Prom. %", True),
            ("MaxCPU_Pct", "Max CPU %", "CPU M\u00e1x. %", True),
            ("Assessment", "Assessment", "Evaluaci\u00f3n", False),
        ], "No VM right-sizing data found.", "No se encontraron datos de ajuste de tama\u00f1o de VMs.")
        sql_table = render_table(sql_rightsizing, [
            ("Name", "Database", "Base de Datos", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("SKU", "SKU", "SKU", False),
            ("MetricType", "Metric", "M\u00e9trica", False),
            ("AvgUsage_Pct", "Avg Usage %", "Uso Prom. %", True),
            ("Assessment", "Assessment", "Evaluaci\u00f3n", False),
        ], "No SQL right-sizing data found.", "No se encontraron datos de ajuste de tama\u00f1o de SQL.")
        plan_table = render_table(plan_rightsizing, [
            ("Name", "App Plan", "Plan de App", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("SKU", "SKU", "SKU", False),
            ("Workers", "Workers", "Workers", True),
            ("AvgCPU_Pct", "Avg CPU %", "CPU Prom. %", True),
            ("Assessment", "Assessment", "Evaluaci\u00f3n", False),
        ], "No App Service Plan right-sizing data found.", "No se encontraron datos de ajuste de Planes de App Service.")
        storage_table = render_table(storage_activity, [
            ("Name", "Storage Account", "Cuenta de Almacenamiento", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("SKU", "SKU", "SKU", False),
            ("AvgDailyTxns", "Avg Daily Txns", "Trans. Diarias Prom.", True),
            ("Assessment", "Assessment", "Evaluaci\u00f3n", False),
        ], "No storage activity data found.", "No se encontraron datos de actividad de almacenamiento.")
        orphaned_table = render_table(orphaned, [
            ("Type", "Type", "Tipo", False),
            ("Name", "Name", "Nombre", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("Detail", "Detail", "Detalle", False),
        ], "No orphaned resources found.", "No se encontraron recursos hu\u00e9rfanos.")
        dealloc_table = render_table(dealloc_vms, [
            ("name", "VM", "VM", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("vmSize", "Size", "Tama\u00f1o", False),
            ("location", "Location", "Ubicaci\u00f3n", False),
        ], "No deallocated VMs found.", "No se encontraron VMs desasignadas.")
        unattached_disks_table = render_table(unattached_disks, [
            ("name", "Disk", "Disco", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("diskSizeGB", "Size (GB)", "Tama\u00f1o (GB)", True),
            ("skuName", "SKU", "SKU", False),
        ], "No unattached disks found.", "No se encontraron discos no adjuntos.")

        return f"""
        <div class="security-summary-band">
            <div>
                <div class="security-eyebrow" style="color: var(--cp-success);">COST OPTIMIZATION</div>
                <h2>{bi('Cost optimization opportunities', 'Oportunidades de optimizaci\u00f3n de costos')}</h2>
                <p>{bi('Right-sizing, idle/orphaned assets, Advisor cost recommendations, and reservation coverage in one view.', 'Ajuste de tama\u00f1o, activos inactivos/hu\u00e9rfanos, recomendaciones de costos de Advisor y cobertura de reservas en una sola vista.')}</p>
            </div>
            <div class="security-score-mark" style="color: var(--cp-success);">{savings_text}</div>
        </div>
        <div class="security-stats">{stat_html}</div>

        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-success);">{bi('Right-sizing opportunities', 'Oportunidades de ajuste de tama\u00f1o')}</div>
            <div class="security-two-column">
                <div class="security-panel"><h3>{bi('Virtual Machines', 'M\u00e1quinas Virtuales')}</h3>{render_distribution(vm_rightsizing)}{vm_table}</div>
                <div class="security-panel"><h3>{bi('SQL Databases', 'Bases de Datos SQL')}</h3>{render_distribution(sql_rightsizing)}{sql_table}</div>
            </div>
            <div class="security-two-column">
                <div class="security-panel"><h3>{bi('App Service Plans', 'Planes de App Service')}</h3>{render_distribution(plan_rightsizing)}{plan_table}</div>
                <div class="security-panel"><h3>{bi('Storage Accounts', 'Cuentas de Almacenamiento')}</h3>{render_distribution(storage_activity)}{storage_table}</div>
            </div>
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-success);">{bi('Idle and orphaned resources', 'Recursos inactivos y hu\u00e9rfanos')}</div>
            <div class="security-two-column">
                <div class="security-panel"><h3>{bi('Deallocated VMs', 'VMs Desasignadas')}</h3>{dealloc_table}</div>
                <div class="security-panel"><h3>{bi('Unattached Managed Disks', 'Discos Administrados No Adjuntos')}</h3>{unattached_disks_table}</div>
            </div>
            <div class="security-panel">
                <h3>{bi('Orphaned resources (all types)', 'Recursos hu\u00e9rfanos (todos los tipos)')}</h3>
                {orphaned_table}
            </div>
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-success);">{bi('Advisor cost recommendations', 'Recomendaciones de costos de Advisor')}</div>
            {cost_recs_table}
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-success);">{bi('Reservation and Savings Plan opportunities', 'Oportunidades de Reservas y Planes de Ahorro')}</div>
            {reservation_table}
        </div>"""

    def _render_actions_for_pillar(self, pillar: str) -> str:
        """Render the action-item cards for a single pillar (used by the dedicated pillar tabs)."""
        items = [a for a in self.engine.action_items if a["pillar"] == pillar]
        if not items:
            return f'<p class="empty-state">{bi("No action items for this pillar.", "Sin elementos de acci\u00f3n para este pilar.")}</p>'
        html = ""
        for i, action in enumerate(items, 1):
            color = DashboardConfig.SEVERITY_COLORS.get(action["severity"], "var(--cp-text-muted)")
            resources_html = ""
            if action.get("resources"):
                resources_html = "<ul class='resource-list'>" + "".join(f"<li>{r}</li>" for r in action["resources"] if r) + "</ul>"
            severity_badge = bi(action['severity'].upper(), SEVERITY_ES.get(action['severity'], action['severity'].upper()))
            title_text = bi(action['title'], action.get('title_es', action['title']))
            desc_text = bi(action['description'], action.get('description_es', action['description']))
            html += f"""
            <div class="action-item" style="border-left: 4px solid {color};">
                <div class="action-header">
                    <span class="action-num">#{i}</span>
                    <span class="severity-badge" style="background:{color};">{severity_badge}</span>
                    <span class="action-title">{title_text}</span>
                </div>
                <p class="action-desc">{desc_text}</p>
                {resources_html}
            </div>"""
        return html

    def _render_subcategory_table(self, pillar: str) -> str:
        """Render the subcategory weight/health breakdown table for a single pillar (used by the dedicated pillar tabs)."""
        rows = self.engine.breakdowns.get(pillar, {}).get("rows", [])
        if not rows:
            return f'<p class="empty-state">{bi("Score breakdown unavailable.", "Desglose de puntaje no disponible.")}</p>'
        rows_html = "".join(
            f"<tr><td>{bi(r['subcategory'], SUBCATEGORY_ES.get(r['subcategory'], r['subcategory']))}</td><td class='num'>{r['weight']}</td>"
            f"<td class='num'>{r['healthy']}/{r['total']}</td><td class='num'>{r['pct']}%</td></tr>"
            for r in rows
        )
        return f"""<div class="security-table-wrap"><table>
            <thead><tr><th>{bi('Subcategory', 'Subcategor\u00eda')}</th><th>{bi('Weight', 'Peso')}</th><th>{bi('Healthy/Total', 'Saludable/Total')}</th><th>{bi('Score', 'Puntaje')}</th></tr></thead>
            <tbody>{rows_html}</tbody>
        </table></div>"""

    def _render_reliability_view(self) -> str:
        """Render reliability posture: Availability Zones, backup coverage, DDoS protection, and Advisor HA recommendations."""
        discovery = self.data.get("discovery", {})
        governance = self.data.get("governance", {})
        advisor = self.data.get("advisor", {})

        def real_rows(source, sheet_name):
            return [
                row for row in source.get(sheet_name, [])
                if not row.get("Result") and str(row.get("DataStatus", "")).lower() != "nodata"
            ]

        def render_table(rows, columns, empty_en, empty_es, limit=15):
            if not rows:
                return f'<p class="empty-state">{bi(empty_en, empty_es)}</p>'
            header = "".join(f"<th>{bi(label_en, label_es)}</th>" for _, label_en, label_es, _ in columns)
            body = ""
            for i, row in enumerate(rows):
                cells = "".join(
                    f'<td class="{"num" if numeric else ""}">{escape(str(row.get(key, "")))}</td>'
                    for key, _, _, numeric in columns
                )
                row_attr = ' class="extra-row" hidden' if i >= limit else ""
                body += f"<tr{row_attr}>{cells}</tr>"
            more = ""
            if len(rows) > limit:
                more_text = bi(f"+ {len(rows) - limit} more", f"+ {len(rows) - limit} m\u00e1s")
                less_text = bi("Show less", "Mostrar menos")
                more = (
                    '<button type="button" class="table-more-note" onclick="toggleTableRows(this)">'
                    f'<span class="more-label">{more_text}</span><span class="less-label">{less_text}</span></button>'
                )
            return f'<div class="security-table-wrap"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>{more}'

        def ddos_enabled(v):
            return str(v.get("ddosProtection", v.get("ddos", ""))).lower() not in ("false", "", "none")

        vms = real_rows(discovery, "VMs")
        vms_no_zone = [v for v in vms if not v.get("zone") and not v.get("availabilityZone")]
        backups = real_rows(discovery, "BackupVaults")
        vnets = real_rows(discovery, "VNets") or real_rows(governance, "VNets")
        vnets_no_ddos = [v for v in vnets if not ddos_enabled(v)]
        reliability_recs = sorted(
            real_rows(advisor, "Reliability"),
            key=lambda r: {"high": 0, "medium": 1, "low": 2}.get(str(r.get("impact", "")).lower(), 3)
        )
        high_impact = [r for r in reliability_recs if str(r.get("impact", "")).lower() == "high"]

        score = self.engine.scores.get("Reliability", 0)
        stats = [
            (f"{score}/100", bi("Reliability Score", "Puntaje de Confiabilidad"), "security-accent"),
            (len(vms_no_zone), bi("VMs without Availability Zone", "VMs sin Zona de Disponibilidad"), "security-warning" if vms_no_zone else ""),
            (len(backups), bi("Backup Vaults Found", "Vaults de Backup Encontrados"), "" if backups else "security-danger"),
            (len(vnets_no_ddos), bi("VNets without DDoS Protection", "VNets sin Protecci\u00f3n DDoS"), "security-warning" if vnets_no_ddos else ""),
            (len(reliability_recs), bi("Advisor Reliability Recs", "Recom. de Confiabilidad"), ""),
            (len(high_impact), bi("High-Impact Recommendations", "Recomendaciones de Alto Impacto"), "security-danger" if high_impact else ""),
        ]
        stat_html = "".join(
            f'<div class="security-stat"><div class="value {color_class}">{value}</div><div class="label">{label}</div></div>'
            for value, label, color_class in stats
        )

        vms_table = render_table(vms_no_zone, [
            ("name", "VM", "VM", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("location", "Location", "Ubicaci\u00f3n", False),
            ("vmSize", "Size", "Tama\u00f1o", False),
            ("powerState", "Power State", "Estado", False),
        ], "All VMs are deployed across Availability Zones.", "Todas las VMs est\u00e1n desplegadas en Zonas de Disponibilidad.")
        backups_table = render_table(backups, [
            ("name", "Backup Vault", "Vault de Backup", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("location", "Location", "Ubicaci\u00f3n", False),
        ], "No Backup/Recovery Services vaults found.", "No se encontraron vaults de Backup/Recovery Services.")
        vnets_table = render_table(vnets_no_ddos, [
            ("name", "VNet", "VNet", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("addressSpace", "Address Space", "Espacio de Direcciones", False),
        ], "All VNets have DDoS Protection enabled.", "Todas las VNets tienen Protecci\u00f3n DDoS habilitada.")
        recs_table = render_table(reliability_recs, [
            ("impact", "Impact", "Impacto", False),
            ("impactedResource", "Resource", "Recurso", False),
            ("problem", "Problem", "Problema", False),
            ("solution", "Solution", "Soluci\u00f3n", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
        ], "No Advisor reliability recommendations found.", "No se encontraron recomendaciones de confiabilidad de Advisor.")

        return f"""
        <div class="security-summary-band">
            <div>
                <div class="security-eyebrow" style="color: var(--cp-link);">RELIABILITY</div>
                <h2>{bi('High availability and disaster recovery', 'Alta disponibilidad y recuperaci\u00f3n ante desastres')}</h2>
                <p>{bi('Availability Zones, backup coverage, DDoS protection, and Advisor high-availability recommendations in one view.', 'Zonas de Disponibilidad, cobertura de backup, protecci\u00f3n DDoS y recomendaciones de alta disponibilidad de Advisor en una sola vista.')}</p>
            </div>
            <div class="security-score-mark" style="color: var(--cp-link);">{score}/100</div>
        </div>
        <div class="security-stats">{stat_html}</div>

        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-link);">{bi('VMs without Availability Zones', 'VMs sin Zonas de Disponibilidad')}</div>
            {vms_table}
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Backup vault coverage', 'Cobertura de vaults de backup')}</h3>{backups_table}</div>
            <div class="security-panel"><h3>{bi('VNets without DDoS Protection', 'VNets sin Protecci\u00f3n DDoS')}</h3>{vnets_table}</div>
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-link);">{bi('Advisor reliability recommendations', 'Recomendaciones de confiabilidad de Advisor')}</div>
            {recs_table}
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-link);">{bi('Score breakdown', 'Desglose de puntaje')}</div>
            {self._render_subcategory_table("Reliability")}
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-link);">{bi('Action items', 'Elementos de acci\u00f3n')}</div>
            {self._render_actions_for_pillar("Reliability")}
        </div>"""

    def _render_performance_view(self) -> str:
        """Render performance efficiency: VM/SQL saturation and Advisor performance recommendations."""
        metrics = self.data.get("metrics", {})
        advisor = self.data.get("advisor", {})

        def real_rows(source, sheet_name):
            return [
                row for row in source.get(sheet_name, [])
                if not row.get("Result") and str(row.get("DataStatus", "")).lower() != "nodata"
            ]

        def render_table(rows, columns, empty_en, empty_es, limit=15):
            if not rows:
                return f'<p class="empty-state">{bi(empty_en, empty_es)}</p>'
            header = "".join(f"<th>{bi(label_en, label_es)}</th>" for _, label_en, label_es, _ in columns)
            body = ""
            for i, row in enumerate(rows):
                cells = "".join(
                    f'<td class="{"num" if numeric else ""}">{escape(str(row.get(key, "")))}</td>'
                    for key, _, _, numeric in columns
                )
                row_attr = ' class="extra-row" hidden' if i >= limit else ""
                body += f"<tr{row_attr}>{cells}</tr>"
            more = ""
            if len(rows) > limit:
                more_text = bi(f"+ {len(rows) - limit} more", f"+ {len(rows) - limit} m\u00e1s")
                less_text = bi("Show less", "Mostrar menos")
                more = (
                    '<button type="button" class="table-more-note" onclick="toggleTableRows(this)">'
                    f'<span class="more-label">{more_text}</span><span class="less-label">{less_text}</span></button>'
                )
            return f'<div class="security-table-wrap"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>{more}'

        vm_metrics = real_rows(metrics, "VM_RightSizing")
        saturated_vm = [v for v in vm_metrics if "saturated" in str(v.get("Assessment", "")).lower()]
        sql_metrics = real_rows(metrics, "SQL_RightSizing")
        saturated_sql = [s for s in sql_metrics if "saturated" in str(s.get("Assessment", "")).lower()]
        perf_recs = real_rows(advisor, "Performance")

        score = self.engine.scores.get("Performance Efficiency", 0)
        stats = [
            (f"{score}/100", bi("Performance Score", "Puntaje de Rendimiento"), "security-accent"),
            (len(saturated_vm), bi("Saturated VMs (>80% CPU)", "VMs Saturadas (>80% CPU)"), "security-danger" if saturated_vm else ""),
            (len(saturated_sql), bi("Saturated SQL DBs", "BD SQL Saturadas"), "security-danger" if saturated_sql else ""),
            (len(perf_recs), bi("Advisor Performance Recs", "Recom. de Rendimiento"), ""),
        ]
        stat_html = "".join(
            f'<div class="security-stat"><div class="value {color_class}">{value}</div><div class="label">{label}</div></div>'
            for value, label, color_class in stats
        )

        vm_table = render_table(saturated_vm, [
            ("Name", "VM", "VM", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("VMSize", "Size", "Tama\u00f1o", False),
            ("AvgCPU_Pct", "Avg CPU %", "CPU Prom. %", True),
            ("MaxCPU_Pct", "Max CPU %", "CPU M\u00e1x. %", True),
        ], "No saturated VMs found.", "No se encontraron VMs saturadas.")
        sql_table = render_table(saturated_sql, [
            ("Name", "Database", "Base de Datos", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("SKU", "SKU", "SKU", False),
            ("MetricType", "Metric", "M\u00e9trica", False),
            ("AvgUsage_Pct", "Avg Usage %", "Uso Prom. %", True),
        ], "No saturated SQL databases found.", "No se encontraron bases de datos SQL saturadas.")
        recs_table = render_table(perf_recs, [
            ("impact", "Impact", "Impacto", False),
            ("impactedResource", "Resource", "Recurso", False),
            ("problem", "Problem", "Problema", False),
            ("solution", "Solution", "Soluci\u00f3n", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
        ], "No Advisor performance recommendations found.", "No se encontraron recomendaciones de rendimiento de Advisor.")

        return f"""
        <div class="security-summary-band">
            <div>
                <div class="security-eyebrow" style="color: var(--cp-warning);">PERFORMANCE EFFICIENCY</div>
                <h2>{bi('Workload scaling and utilization', 'Escalado y utilizaci\u00f3n de cargas de trabajo')}</h2>
                <p>{bi('Saturated VMs and SQL databases, plus Advisor performance recommendations in one view.', 'VMs y bases de datos SQL saturadas, m\u00e1s recomendaciones de rendimiento de Advisor en una sola vista.')}</p>
            </div>
            <div class="security-score-mark" style="color: var(--cp-warning);">{score}/100</div>
        </div>
        <div class="security-stats">{stat_html}</div>

        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-warning);">{bi('Resource saturation', 'Saturaci\u00f3n de recursos')}</div>
            <div class="security-two-column">
                <div class="security-panel"><h3>{bi('Virtual Machines', 'M\u00e1quinas Virtuales')}</h3>{vm_table}</div>
                <div class="security-panel"><h3>{bi('SQL Databases', 'Bases de Datos SQL')}</h3>{sql_table}</div>
            </div>
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-warning);">{bi('Advisor performance recommendations', 'Recomendaciones de rendimiento de Advisor')}</div>
            {recs_table}
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-warning);">{bi('Score breakdown', 'Desglose de puntaje')}</div>
            {self._render_subcategory_table("Performance Efficiency")}
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-warning);">{bi('Action items', 'Elementos de acci\u00f3n')}</div>
            {self._render_actions_for_pillar("Performance Efficiency")}
        </div>"""

    def _render_operational_view(self) -> str:
        """Render operational excellence: diagnostics coverage, tagging, resource locks, policy compliance, and Advisor recommendations."""
        metrics = self.data.get("metrics", {})
        discovery = self.data.get("discovery", {})
        governance = self.data.get("governance", {})
        advisor = self.data.get("advisor", {})

        def real_rows(source, sheet_name):
            return [
                row for row in source.get(sheet_name, [])
                if not row.get("Result") and str(row.get("DataStatus", "")).lower() != "nodata"
            ]

        def as_int(value):
            try:
                return int(value or 0)
            except (TypeError, ValueError):
                return 0

        def render_table(rows, columns, empty_en, empty_es, limit=15):
            if not rows:
                return f'<p class="empty-state">{bi(empty_en, empty_es)}</p>'
            header = "".join(f"<th>{bi(label_en, label_es)}</th>" for _, label_en, label_es, _ in columns)
            body = ""
            for i, row in enumerate(rows):
                cells = "".join(
                    f'<td class="{"num" if numeric else ""}">{escape(str(row.get(key, "")))}</td>'
                    for key, _, _, numeric in columns
                )
                row_attr = ' class="extra-row" hidden' if i >= limit else ""
                body += f"<tr{row_attr}>{cells}</tr>"
            more = ""
            if len(rows) > limit:
                more_text = bi(f"+ {len(rows) - limit} more", f"+ {len(rows) - limit} m\u00e1s")
                less_text = bi("Show less", "Mostrar menos")
                more = (
                    '<button type="button" class="table-more-note" onclick="toggleTableRows(this)">'
                    f'<span class="more-label">{more_text}</span><span class="less-label">{less_text}</span></button>'
                )
            return f'<div class="security-table-wrap"><table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table></div>{more}'

        diag = real_rows(metrics, "DiagnosticsCoverage")
        no_diag = [d for d in diag if str(d.get("HasDiagnostics", d.get("Gap", ""))).lower() in ("false", "no diagnostics configured")]

        TAGGABLE_SHEETS = ["VMs", "AppServices", "AKS", "VNets", "NSGs", "LoadBalancers",
                           "Firewalls", "PublicIPs", "Storage", "Databases", "KeyVaults"]
        untagged = []
        total_taggable = 0
        for sheet in TAGGABLE_SHEETS:
            for r in real_rows(discovery, sheet):
                total_taggable += 1
                tags = r.get("tags")
                if not tags or str(tags).strip().lower() in ("", "none", "{}"):
                    untagged.append({
                        "Type": sheet,
                        "Name": r.get("name", r.get("Name", "")),
                        "ResourceGroup": r.get("resourceGroup", r.get("ResourceGroup", "")),
                    })

        locks = real_rows(governance, "Locks")
        policy_compliance = real_rows(governance, "PolicyCompliance")
        non_compliant = [c for c in policy_compliance if as_int(c.get("NonCompliantCount", 0)) > 0]
        opex_recs = real_rows(advisor, "OperationalExcellence")

        score = self.engine.scores.get("Operational Excellence", 0)
        stats = [
            (f"{score}/100", bi("OpEx Score", "Puntaje OpEx"), "security-accent"),
            (len(no_diag), bi("Missing Diagnostic Settings", "Sin Configuraci\u00f3n de Diagn\u00f3stico"), "security-warning" if no_diag else ""),
            (f"{len(untagged)}/{total_taggable}", bi("Untagged Resources", "Recursos sin Etiquetas"), "security-warning" if untagged else ""),
            (len(locks), bi("Resource Locks", "Bloqueos de Recursos"), "" if locks else "security-danger"),
            (len(non_compliant), bi("Non-Compliant Policy Evaluations", "Evaluaciones No Conformes"), "security-warning" if non_compliant else ""),
            (len(opex_recs), bi("Advisor OpEx Recs", "Recom. OpEx"), ""),
        ]
        stat_html = "".join(
            f'<div class="security-stat"><div class="value {color_class}">{value}</div><div class="label">{label}</div></div>'
            for value, label, color_class in stats
        )

        diag_table = render_table(no_diag, [
            ("Name", "Resource", "Recurso", False),
            ("Type", "Type", "Tipo", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
            ("Gap", "Gap", "Brecha", False),
        ], "All critical resources have diagnostic settings configured.", "Todos los recursos cr\u00edticos tienen configuraci\u00f3n de diagn\u00f3stico.")
        tags_table = render_table(untagged, [
            ("Type", "Resource Type", "Tipo de Recurso", False),
            ("Name", "Name", "Nombre", False),
            ("ResourceGroup", "Resource Group", "Grupo de Recursos", False),
        ], "All resources are tagged.", "Todos los recursos tienen etiquetas.")
        locks_table = render_table(locks, [
            ("name", "Name", "Nombre", False),
            ("lockLevel", "Level", "Nivel", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
        ], "No resource locks configured.", "No hay bloqueos de recursos configurados.")
        recs_table = render_table(opex_recs, [
            ("impact", "Impact", "Impacto", False),
            ("impactedResource", "Resource", "Recurso", False),
            ("problem", "Problem", "Problema", False),
            ("solution", "Solution", "Soluci\u00f3n", False),
            ("resourceGroup", "Resource Group", "Grupo de Recursos", False),
        ], "No Advisor operational excellence recommendations found.", "No se encontraron recomendaciones de excelencia operativa de Advisor.")

        return f"""
        <div class="security-summary-band">
            <div>
                <div class="security-eyebrow" style="color: var(--cp-accent);">OPERATIONAL EXCELLENCE</div>
                <h2>{bi('Observability, tagging, and governance hygiene', 'Observabilidad, etiquetado e higiene de gobernanza')}</h2>
                <p>{bi('Diagnostic settings coverage, tagging compliance, resource locks, and Advisor operational recommendations in one view.', 'Cobertura de configuraci\u00f3n de diagn\u00f3stico, cumplimiento de etiquetado, bloqueos de recursos y recomendaciones operativas de Advisor en una sola vista.')}</p>
            </div>
            <div class="security-score-mark" style="color: var(--cp-accent);">{score}/100</div>
        </div>
        <div class="security-stats">{stat_html}</div>

        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-accent);">{bi('Diagnostic settings coverage', 'Cobertura de configuraci\u00f3n de diagn\u00f3stico')}</div>
            {diag_table}
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-accent);">{bi('Untagged resources', 'Recursos sin etiquetas')}</div>
            {tags_table}
        </div>
        <div class="security-two-column">
            <div class="security-panel"><h3>{bi('Resource locks', 'Bloqueos de recursos')}</h3>{locks_table}</div>
            <div class="security-panel"><h3>{bi('Advisor operational excellence recommendations', 'Recomendaciones de excelencia operativa de Advisor')}</h3>{recs_table}</div>
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-accent);">{bi('Score breakdown', 'Desglose de puntaje')}</div>
            {self._render_subcategory_table("Operational Excellence")}
        </div>
        <div class="section">
            <div class="section-title" style="border-bottom-color: var(--cp-accent);">{bi('Action items', 'Elementos de acci\u00f3n')}</div>
            {self._render_actions_for_pillar("Operational Excellence")}
        </div>"""

    def _render_score_breakdown(self) -> str:
        """Render the 'why' behind each pillar score per the Advisor Score model."""
        b = self.engine.breakdowns
        sections = []
        pillar_label = lambda p: bi(p, PILLAR_ES.get(p, p))
        insufficient_note = bi(
            "Insufficient data — no discovered resource inventory (01_Discovery/ResourceSummary) to compute this pillar's score.",
            "Datos insuficientes: no hay inventario de recursos descubiertos (01_Discovery/ResourceSummary) para calcular el puntaje de este pilar."
        )

        def subcategory_card(pillar):
            info = b.get(pillar, {})
            if not info.get("available"):
                return f"""<div class="breakdown-card">
                <h4>{pillar_label(pillar)}</h4>
                <p class="breakdown-note">{insufficient_note}</p>
            </div>"""
            rows_html = "".join(
                f"<tr><td>{bi(r['subcategory'], SUBCATEGORY_ES.get(r['subcategory'], r['subcategory']))}</td><td class='num'>{r['weight']}</td>"
                f"<td class='num'>{r['healthy']}/{r['total']}</td><td class='num'>{r['pct']}%</td></tr>"
                for r in info["rows"]
            )
            return f"""<div class="breakdown-card">
                <h4>{pillar_label(pillar)} <span class="breakdown-score">{self.engine.scores[pillar]}/100</span></h4>
                <table>
                    <thead><tr><th>{bi('Subcategory', 'Subcategor\u00eda')}</th><th>{bi('Weight', 'Peso')}</th><th>{bi('Healthy/Total', 'Saludable/Total')}</th><th>{bi('Score', 'Puntaje')}</th></tr></thead>
                    <tbody>{rows_html}</tbody>
                </table>
            </div>"""

        sections.append(subcategory_card("Reliability"))
        sections.append(subcategory_card("Performance Efficiency"))
        sections.append(subcategory_card("Operational Excellence"))

        security = b.get("Security", {})
        if security.get("available"):
            findings_html = "".join(
                f"<tr><td>{bi(label_en, label_es)}</td><td class='num'>{count}</td></tr>"
                for label_en, label_es, count in security.get("findings", [])
            )
            note = bi(
                f"Microsoft Defender Secure Score average across {security.get('subscriptions_scored', 0)} subscription(s): <strong>{security.get('secure_score_pct')}%</strong>.",
                f"Promedio del Secure Score de Microsoft Defender en {security.get('subscriptions_scored', 0)} suscripci\u00f3n(es): <strong>{security.get('secure_score_pct')}%</strong>."
            )
            sections.append(f"""<div class="breakdown-card">
                <h4>{pillar_label('Security')} <span class="breakdown-score">{self.engine.scores['Security']}/100</span></h4>
                <p class="breakdown-note">{note}</p>
                <table>
                    <thead><tr><th>{bi('Finding Category', 'Categor\u00eda de Hallazgo')}</th><th>{bi('Count', 'Cantidad')}</th></tr></thead>
                    <tbody>{findings_html}</tbody>
                </table>
            </div>""")
        else:
            note = bi(
                "Insufficient data — no Defender Secure Score records found in 04_Governance/SecureScores.",
                "Datos insuficientes: no se encontraron registros de Secure Score de Defender en 04_Governance/SecureScores."
            )
            sections.append(f"""<div class="breakdown-card">
                <h4>{pillar_label('Security')}</h4>
                <p class="breakdown-note">{note}</p>
            </div>""")

        cost = b.get("Cost Optimization", {})
        if cost.get("available"):
            source_rows = []
            for source in cost.get("sources", []):
                name = source["name"]
                affected = min(source["affected"], cost["total"])
                source_healthy = max(0, cost["total"] - affected)
                source_score = round((source_healthy / cost["total"]) * 100)
                source_rows.append(
                    f"<tr><td>{bi(name, COST_SOURCE_ES.get(name, name))}</td>"
                    f"<td class='num'>{source['count']} / {affected}</td>"
                    f"<td class='num'>{source_healthy}/{cost['total']}</td>"
                    f"<td class='num'>{source_score}%</td></tr>"
                )
            combined_label = bi("Combined (deduplicated)", "Combinado (sin duplicados)")
            combined_findings = sum(source["count"] for source in cost.get("sources", []))
            source_rows.append(
                f"<tr><td><strong>{combined_label}</strong></td>"
                f"<td class='num'><strong>{combined_findings} / {cost['impacted']}</strong></td>"
                f"<td class='num'><strong>{cost['healthy']}/{cost['total']}</strong></td>"
                f"<td class='num'><strong>{self.engine.scores['Cost Optimization']}%</strong></td></tr>"
            )
            sources_html = "".join(source_rows)
            note = bi(
                "Findings/Affected shows finding count first and unique affected scopes or resources second. The combined row deduplicates identifiers.",
                "Hallazgos/Afectados muestra primero los hallazgos y luego los ámbitos o recursos únicos afectados. La fila combinada elimina duplicados."
            )
            sections.append(f"""<div class="breakdown-card cost-breakdown-card">
                <h4>{pillar_label('Cost Optimization')} <span class="breakdown-score">{self.engine.scores['Cost Optimization']}/100</span></h4>
                <p class="breakdown-note">{note}</p>
                <div class="breakdown-table-wrap">
                    <table>
                        <colgroup><col class="cost-source-col"><col class="cost-findings-col"><col><col></colgroup>
                        <thead><tr><th>{bi('Finding Source', 'Fuente del Hallazgo')}</th><th>{bi('Findings/Affected', 'Hallazgos/Afectados')}</th><th>{bi('Healthy/Total', 'Saludable/Total')}</th><th>{bi('Score', 'Puntaje')}</th></tr></thead>
                        <tbody>{sources_html}</tbody>
                    </table>
                </div>
            </div>""")
        else:
            note = bi(
                "Insufficient data — no discovered resource inventory to compute this pillar's score.",
                "Datos insuficientes: no hay inventario de recursos descubiertos para calcular el puntaje de este pilar."
            )
            sections.append(f"""<div class="breakdown-card">
                <h4>{pillar_label('Cost Optimization')}</h4>
                <p class="breakdown-note">{note}</p>
            </div>""")

        return "".join(sections)
    
    def generate(self) -> str:
        overall = self.engine.get_overall_score()
        scores = self.engine.scores
        score_breakdown_html = self._render_score_breakdown()
        governance_view_html = self._render_governance_view()
        security_view_html = self._render_security_view()
        cost_view_html = self._render_cost_view()
        reliability_view_html = self._render_reliability_view()
        performance_view_html = self._render_performance_view()
        operational_view_html = self._render_operational_view()
        action_items = sorted(self.engine.action_items, 
                            key=lambda x: {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}.get(x["severity"], 5))
        
        # Count stats
        total_resources = sum(int(r.get("Count", 0) or 0) for r in self.data.get("discovery", {}).get("ResourceSummary", []))
        total_subs = len(self.data.get("discovery", {}).get("Subscriptions", []) or self.data.get("governance", {}).get("Subscriptions", []))
        total_actions = len(action_items)
        critical_actions = len([a for a in action_items if a["severity"] in ("critical", "high")])
        
        # Build action items HTML
        actions_html = ""
        for i, action in enumerate(action_items, 1):
            color = DashboardConfig.SEVERITY_COLORS.get(action["severity"], "var(--cp-text-muted)")
            resources_html = ""
            if action.get("resources"):
                resources_html = "<ul class='resource-list'>" + "".join(f"<li>{r}</li>" for r in action["resources"] if r) + "</ul>"
            
            severity_badge = bi(action['severity'].upper(), SEVERITY_ES.get(action['severity'], action['severity'].upper()))
            pillar_tag = bi(action['pillar'], PILLAR_ES.get(action['pillar'], action['pillar']))
            title_text = bi(action['title'], action.get('title_es', action['title']))
            desc_text = bi(action['description'], action.get('description_es', action['description']))
            actions_html += f"""
            <div class="action-item" style="border-left: 4px solid {color};">
                <div class="action-header">
                    <span class="action-num">#{i}</span>
                    <span class="severity-badge" style="background:{color};">{severity_badge}</span>
                    <span class="pillar-tag">{pillar_tag}</span>
                    <span class="action-title">{title_text}</span>
                </div>
                <p class="action-desc">{desc_text}</p>
                {resources_html}
            </div>"""
        
        # Build pillar score cards
        pillar_cards = ""
        for pillar, sc in scores.items():
            color = DashboardConfig.PILLAR_COLORS[pillar]
            status_icon = "✅" if sc >= 80 else "⚠️" if sc >= 50 else "🔴"
            pillar_cards += f"""
            <div class="pillar-card">
                <div class="pillar-score" style="color:{color};">{sc}</div>
                <div class="pillar-bar">
                    <div class="pillar-bar-fill" style="width:{sc}%; background:{color};"></div>
                </div>
                <div class="pillar-name">{status_icon} {bi(pillar, PILLAR_ES.get(pillar, pillar))}</div>
            </div>"""
        
        # Build resource summary table
        resource_summary = self.data.get("discovery", {}).get("ResourceSummary", [])
        top_resources = sorted(resource_summary, key=lambda x: int(x.get("Count", 0) or 0), reverse=True)[:20]
        resource_rows = ""
        for r in top_resources:
            rtype = str(r.get("type", "")).split("/")[-1] if "/" in str(r.get("type", "")) else r.get("type", "")
            resource_rows += f"<tr><td>{r.get('type', '')}</td><td>{r.get('location', '')}</td><td class='num'>{r.get('Count', 0)}</td></tr>"
        
        # Advisor summary
        advisor_summary = self.data.get("advisor", {}).get("SummaryByCategory", [])
        advisor_rows = ""
        for a in advisor_summary:
            advisor_rows += f"<tr><td>{a.get('category', '')}</td><td>{a.get('impact', '')}</td><td class='num'>{a.get('Count', 0)}</td></tr>"
        
        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Azure WAF/CAF Workshop - Consolidated Dashboard</title>
<script>
    (() => {{
        const param = new URLSearchParams(window.location.search).get("scoutTheme");
        let saved = null;
        try {{ saved = localStorage.getItem("wafDashboardTheme"); }} catch (e) {{}}
        const theme = ["light", "dark"].includes(param)
            ? param
            : (["light", "dark"].includes(saved) ? saved : "light");
        document.documentElement.setAttribute("data-theme", theme);
    }})();
</script>
<style>
:root {{
    color-scheme: light;
    --cp-bg: #f5f5f7;
    --cp-bg-elevated: #ffffff;
    --cp-surface: #ffffff;
    --cp-surface-soft: #f5f5f7;
    --cp-border: rgba(0, 0, 0, 0.08);
    --cp-border-strong: rgba(0, 0, 0, 0.16);
    --cp-text: #1d1d1f;
    --cp-text-muted: #6e6e73;
    --cp-text-soft: #86868b;
    --cp-accent: #0071e3;
    --cp-accent-hover: #0077ed;
    --cp-accent-soft: rgba(0, 113, 227, 0.08);
    --cp-accent-fg: #ffffff;
    --cp-success: #1d9a4a;
    --cp-danger: #ff3b30;
    --cp-warning: #ff9500;
    --cp-link: #0071e3;
    --cp-shadow: 0 2px 8px rgba(0, 0, 0, 0.04), 0 12px 32px rgba(0, 0, 0, 0.06);
    --cp-overlay: rgba(255, 255, 255, 0.75);
    --cp-panel: rgba(255, 255, 255, 0.72);
    --cp-panel-strong: rgba(255, 255, 255, 0.94);
    --cp-sheen: rgba(255, 255, 255, 0.6);
    --cp-highlight: rgba(0, 113, 227, 0.1);
    --overview-header-start: #f5f5f7;
    --overview-header-end: #f5f5f7;
    --cp-radius-lg: 22px;
    --cp-radius-md: 16px;
    --cp-radius-sm: 12px;
    --cp-ease: cubic-bezier(0.25, 0.1, 0.25, 1);
}}
html[data-theme="dark"] {{
    color-scheme: dark;
    --cp-bg: #000000;
    --cp-bg-elevated: #1c1c1e;
    --cp-surface: #1c1c1e;
    --cp-surface-soft: #2c2c2e;
    --cp-border: rgba(255, 255, 255, 0.08);
    --cp-border-strong: rgba(255, 255, 255, 0.16);
    --cp-text: #f5f5f7;
    --cp-text-muted: #a1a1a6;
    --cp-text-soft: #86868b;
    --cp-accent: #0a84ff;
    --cp-accent-hover: #409cff;
    --cp-accent-soft: rgba(10, 132, 255, 0.16);
    --cp-accent-fg: #ffffff;
    --cp-success: #30d158;
    --cp-danger: #ff453a;
    --cp-warning: #ff9f0a;
    --cp-link: #0a84ff;
    --cp-shadow: 0 2px 8px rgba(0, 0, 0, 0.3), 0 12px 32px rgba(0, 0, 0, 0.4);
    --cp-overlay: rgba(28, 28, 30, 0.75);
    --cp-panel: rgba(28, 28, 30, 0.72);
    --cp-panel-strong: rgba(28, 28, 30, 0.94);
    --cp-sheen: rgba(255, 255, 255, 0.04);
    --cp-highlight: rgba(10, 132, 255, 0.16);
    --overview-header-start: #1c1c1e;
    --overview-header-end: #1c1c1e;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
        font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display", "SF Pro Text", "Segoe UI", Helvetica, Arial, sans-serif;
        background: var(--cp-bg); color: var(--cp-text);
    line-height: 1.47059; -webkit-font-smoothing: antialiased; letter-spacing: -0.01em;
}}
.dashboard {{ max-width: 1400px; margin: 0 auto; padding: 32px 24px 24px; }}

/* Header */
.header {{
    background: linear-gradient(135deg, #063466 0%, #00152e 100%);
    color: #ffffff; padding: 40px 40px 32px; border-radius: var(--cp-radius-lg);
    border: 1px solid rgba(255, 255, 255, 0.08); box-shadow: var(--cp-shadow);
    margin-bottom: 24px; position: relative;
}}
.header h1 {{ font-size: 30px; font-weight: 700; letter-spacing: -0.02em; margin: 0; min-width: 0; }}
.header-brand {{ display: flex; align-items: center; gap: 24px; margin-bottom: 12px; }}
.header-logo {{ width: 96px; height: 96px; flex-shrink: 0; }}
.header-brand .az-icon.header-logo {{ width: 96px; height: 96px; }}
.header-title-sub {{ display: block; font-size: 22px; font-weight: 700; margin-top: 2px; }}
.header .subtitle {{ color: rgba(255, 255, 255, 0.78); font-size: 14px; font-weight: 400; }}

/* Dashboard views (segmented control) */
.view-tabs {{
    display: flex; flex-wrap: wrap; justify-content: space-evenly; width: 100%; gap: 10px; padding: 4px;
    margin: 0 auto 24px; background: var(--cp-surface-soft); border: 1px solid var(--cp-border); border-radius: 980px;
    position: relative;
}}
.view-tab-thumb {{
    position: absolute; top: 4px; left: 0; height: calc(100% - 8px); width: 0;
    background: var(--cp-surface); border-radius: 980px; box-shadow: 0 1px 4px rgba(0,0,0,0.12);
    transition: transform 0.4s var(--cp-ease), width 0.4s var(--cp-ease);
    will-change: transform, width; z-index: 0; pointer-events: none;
}}
.view-tab {{
    border: 0; background: transparent; color: var(--cp-text-muted); cursor: pointer;
    padding: 8px 20px; border-radius: 980px; font: inherit; font-size: 13px; font-weight: 590;
    transition: color 0.25s var(--cp-ease); position: relative; z-index: 1;
}}
.view-tab[aria-selected="true"] {{ color: var(--cp-text); }}
.dashboard-view {{ transition: opacity 0.22s var(--cp-ease), transform 0.22s var(--cp-ease); opacity: 1; transform: translateY(0); }}
.dashboard-view.view-fade-out {{ opacity: 0; transform: translateY(6px); }}
.dashboard-view.view-fade-in {{ opacity: 0; transform: translateY(6px); }}
.dashboard-view[hidden] {{ display: none; }}

/* Dedicated security view */
.security-summary-band {{
    display: flex; justify-content: space-between; align-items: center; gap: 24px;
    background: var(--cp-surface); color: var(--cp-text); padding: 28px 32px;
    border: 1px solid var(--cp-border); border-radius: var(--cp-radius-lg); box-shadow: var(--cp-shadow);
    margin-bottom: 18px;
}}
.security-summary-band h2 {{ font-size: 26px; font-weight: 600; letter-spacing: -0.02em; margin: 2px 0 6px; }}
.security-summary-band p {{ color: var(--cp-text-muted); font-size: 13px; }}
.security-eyebrow {{ color: var(--cp-danger); font-size: 10px; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; }}
.security-score-mark {{ font-size: 44px; font-weight: 700; letter-spacing: -0.02em; color: var(--cp-danger); white-space: nowrap; }}
.security-stats {{ display: grid; grid-template-columns: repeat(6, minmax(0, 1fr)); gap: 12px; margin-bottom: 24px; }}
.security-stat {{
    background: var(--cp-surface); border: 1px solid var(--cp-border); border-radius: var(--cp-radius-md);
    padding: 18px; min-width: 0; box-shadow: var(--cp-shadow); transition: transform 0.2s var(--cp-ease);
}}
.security-stat:hover {{ transform: translateY(-2px); }}
.security-stat .value {{ font-size: 27px; font-weight: 700; letter-spacing: -0.02em; }}
.security-stat .label {{ color: var(--cp-text-muted); font-size: 11px; margin-top: 4px; }}
.security-accent {{ color: var(--cp-link); }}
.security-danger {{ color: var(--cp-danger); }}
.security-warning {{ color: var(--cp-warning); }}
.security-title {{ border-bottom-color: var(--cp-danger); }}
.security-two-column {{ display: grid; grid-template-columns: minmax(0, 1fr) minmax(0, 1fr); gap: 18px; margin-bottom: 24px; }}
.security-panel {{
    background: var(--cp-surface); padding: 20px; min-width: 0;
    border: 1px solid var(--cp-border); border-radius: var(--cp-radius-md); box-shadow: var(--cp-shadow);
}}
.security-panel h3 {{ font-size: 14px; font-weight: 590; margin-bottom: 12px; }}
.security-table-wrap {{ width: 100%; overflow-x: auto; }}
.security-table-wrap table {{ min-width: 540px; }}
.empty-state {{ color: var(--cp-text-muted); font-size: 12px; padding: 16px 0; }}
.source-status {{ display: inline-block; padding: 3px 10px; border-radius: 980px; font-size: 10px; font-weight: 700; }}
.source-available, .source-nodata {{ background: rgba(48, 209, 88, 0.12); color: var(--cp-success); }}
.source-partial {{ background: rgba(255, 149, 0, 0.14); color: var(--cp-warning); }}
.source-forbidden, .source-unavailable, .source-error {{ background: rgba(255, 59, 48, 0.12); color: var(--cp-danger); }}
.source-skipped {{ background: var(--cp-surface-soft); color: var(--cp-text-muted); }}
.security-success {{ color: var(--cp-success); }}
.cost-dist {{ display: flex; flex-direction: column; gap: 6px; margin-bottom: 14px; }}
.cost-dist-row {{ display: grid; grid-template-columns: minmax(90px, auto) 1fr 24px; align-items: center; gap: 8px; font-size: 11px; }}
.cost-dist-label {{ color: var(--cp-text-muted); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}
.cost-dist-track {{ height: 6px; border-radius: 980px; background: var(--cp-surface-soft); overflow: hidden; }}
.cost-dist-fill {{ height: 100%; border-radius: 980px; }}
.cost-dist-count {{ text-align: right; font-weight: 600; }}
@media (max-width: 1050px) {{ .security-stats {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }} }}
@media (max-width: 760px) {{
    .view-tabs {{ width: 100%; }}
    .view-tab {{ flex: 1; }}
    .security-summary-band {{ align-items: flex-start; padding: 22px 18px; }}
    .security-score-mark {{ font-size: 30px; }}
    .security-stats {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    .security-two-column {{ grid-template-columns: minmax(0, 1fr); }}
}}

/* Overall Score */
.score-section {{
    display: grid; grid-template-columns: 200px 1fr;
    gap: 24px; margin-bottom: 24px;
    background: var(--cp-surface); border-radius: var(--cp-radius-lg); padding: 32px;
    border: 1px solid var(--cp-border); box-shadow: var(--cp-shadow);
}}
.overall-score {{
    display: flex; flex-direction: column; align-items: center; justify-content: center;
}}
.score-circle {{
    width: 140px; height: 140px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 48px; font-weight: 700;
    background: conic-gradient(
        {'var(--cp-success)' if overall >= 80 else 'var(--cp-warning)' if overall >= 50 else 'var(--cp-danger)'} {overall * 3.6}deg,
        var(--cp-border) {overall * 3.6}deg
    );
    position: relative;
}}
.score-circle::after {{
    content: '{overall}'; position: absolute;
    width: 114px; height: 114px; border-radius: 50%;
    background: var(--cp-surface); display: flex; align-items: center; justify-content: center;
    font-size: 40px; font-weight: 700; letter-spacing: -0.02em;
    color: {'var(--cp-success)' if overall >= 80 else 'var(--cp-warning)' if overall >= 50 else 'var(--cp-danger)'};
}}
.score-label {{ margin-top: 12px; font-size: 13px; color: var(--cp-text-muted); font-weight: 590; }}

.pillars-grid {{
    display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px;
}}
.pillar-card {{
    background: var(--cp-surface-soft); border-radius: var(--cp-radius-md); padding: 18px;
    border: 1px solid var(--cp-border); transition: transform 0.2s var(--cp-ease), box-shadow 0.2s var(--cp-ease);
}}
.pillar-card:hover {{ transform: translateY(-2px); box-shadow: var(--cp-shadow); }}
.pillar-score {{ font-size: 32px; font-weight: 700; letter-spacing: -0.02em; }}
.pillar-bar {{ height: 5px; background: var(--cp-border); border-radius: 980px; margin: 10px 0; overflow: hidden; }}
.pillar-bar-fill {{ height: 100%; border-radius: 980px; transition: width 0.6s var(--cp-ease); }}
.pillar-name {{ font-size: 12px; font-weight: 590; color: var(--cp-text-muted); }}

@media (max-width: 700px) {{
    .dashboard {{ padding: 16px; }}
    .header {{ padding: 72px 20px 24px; }}
    .header h1 {{ padding-right: 0; font-size: 24px; }}
    .header-title-sub {{ font-size: 18px; }}
    .header-logo {{ width: 40px; height: 40px; }}
    .header-brand .az-icon.header-logo {{ width: 40px; height: 40px; }}
    .header-controls {{ top: 16px; right: 16px; }}
    .score-section {{ grid-template-columns: minmax(0, 1fr); padding: 20px; }}
    .pillars-grid {{ grid-template-columns: repeat(auto-fit, minmax(min(200px, 100%), 1fr)); min-width: 0; }}
    .pillar-card {{ min-width: 0; }}
}}

/* Score Breakdown */
.methodology-note {{
    background: var(--cp-accent-soft); border: 1px solid var(--cp-border); border-radius: var(--cp-radius-md);
    padding: 14px 18px; font-size: 12px; color: var(--cp-text); margin-bottom: 16px;
}}
.methodology-note a {{ color: var(--cp-link); }}
.breakdown-grid {{
    display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 16px;
}}
@media (min-width: 701px) {{ .breakdown-grid {{ grid-auto-rows: 1fr; }} }}
.breakdown-card {{
    background: var(--cp-surface); border-radius: var(--cp-radius-md); padding: 18px;
    border: 1px solid var(--cp-border); box-shadow: var(--cp-shadow); min-width: 0;
    transition: transform 0.2s var(--cp-ease);
}}
.breakdown-card:hover {{ transform: translateY(-2px); }}
.breakdown-card h4 {{ font-size: 13px; font-weight: 590; margin-bottom: 10px; display: flex; justify-content: space-between; }}
.breakdown-score {{ color: var(--cp-link); }}
.breakdown-note {{ font-size: 12px; color: var(--cp-text-muted); }}
.breakdown-table-wrap {{ width: 100%; overflow-x: auto; }}
.cost-breakdown-card table {{ table-layout: fixed; }}
.cost-breakdown-card th,
.cost-breakdown-card td {{ overflow-wrap: anywhere; white-space: normal; }}
.cost-source-col {{ width: 34%; }}
.cost-findings-col {{ width: 26%; }}

/* Governance view (native hierarchy tree, scope insights) */
.mg-tree, .mg-tree ul {{ list-style: none; padding-left: 30px; }}
.mg-tree {{ padding-left: 0; }}
.mg-tree li {{ margin: 8px 0; }}
.mg-node-details > summary {{ list-style: none; cursor: pointer; }}
.mg-node-details > summary::-webkit-details-marker {{ display: none; }}
.mg-node {{
    display: inline-flex; align-items: center; gap: 12px; background: var(--cp-surface);
    border: 1px solid var(--cp-border); border-radius: 980px; padding: 10px 18px; font-size: 15px; font-weight: 590;
}}
.mg-name {{ white-space: nowrap; display: inline-flex; align-items: center; gap: 8px; }}
.az-icon {{ width: 22px; height: 22px; flex-shrink: 0; vertical-align: middle; }}
.mg-badge .az-icon {{ width: 14px; height: 14px; }}
.mg-sub-leaf .az-icon {{ width: 18px; height: 18px; }}
.mg-badge {{
    display: inline-flex; align-items: center; justify-content: center; gap: 4px; min-width: 26px; height: 26px;
    border-radius: 980px; font-size: 12px; font-weight: 700; padding: 0 10px;
}}
.mg-badge-policy {{ background: rgba(0, 120, 212, 0.14); color: var(--cp-link); }}
.mg-badge-rbac {{ background: rgba(255, 149, 0, 0.14); color: var(--cp-warning); }}
.mg-badge-subs {{ background: rgba(48, 209, 88, 0.12); color: var(--cp-success); }}
.mg-sub-leaf {{ font-size: 14px; color: var(--cp-text-muted); padding: 6px 0; list-style: none; display: flex; align-items: center; gap: 8px; }}
.scope-insights {{ display: flex; flex-direction: column; gap: 6px; }}
.scope-detail {{
    background: var(--cp-surface); border: 1px solid var(--cp-border); border-radius: var(--cp-radius-md); padding: 10px 14px;
}}
.scope-detail summary {{ cursor: pointer; font-size: 12px; display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }}
.scope-detail .scope-id {{ color: var(--cp-text-muted); font-weight: 400; }}
.scope-detail p {{ font-size: 12px; margin: 8px 0; }}
.mini-table {{ width: 100%; border-collapse: collapse; font-size: 11px; margin-top: 6px; }}
.mini-table th, .mini-table td {{ border: 1px solid var(--cp-border); padding: 4px 8px; text-align: left; }}
.table-more-note {{
    display: block; width: 100%; text-align: left; font-size: 11px; color: var(--cp-accent); font-weight: 600;
    margin-top: 6px; padding: 4px 0; background: none; border: none; cursor: pointer; font-family: inherit;
}}
.table-more-note:hover {{ text-decoration: underline; }}
.table-more-note .less-label {{ display: none; }}
.table-more-note.expanded .more-label {{ display: none; }}
.table-more-note.expanded .less-label {{ display: inline; }}

/* Stats */
.stats-row {{
    display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 16px; margin-bottom: 24px;
}}
.stat-box {{
    background: var(--cp-surface); border-radius: var(--cp-radius-md); padding: 24px;
    border: 1px solid var(--cp-border); box-shadow: var(--cp-shadow); text-align: center;
    transition: transform 0.2s var(--cp-ease);
}}
.stat-box:hover {{ transform: translateY(-2px); }}
.stat-box .value {{ font-size: 36px; font-weight: 700; letter-spacing: -0.02em; color: var(--cp-link); }}
.stat-box .label {{ font-size: 12px; color: var(--cp-text-muted); margin-top: 4px; }}

/* Action Items */
.section {{ margin-bottom: 24px; }}
.section-title {{
    font-size: 19px; font-weight: 600; letter-spacing: -0.01em; margin-bottom: 16px;
    padding-bottom: 10px; border-bottom: 2px solid var(--cp-accent);
}}
.action-item {{
    background: var(--cp-surface); border-radius: var(--cp-radius-md); padding: 18px; margin-bottom: 12px;
    border: 1px solid var(--cp-border); box-shadow: var(--cp-shadow);
    transition: transform 0.2s var(--cp-ease);
}}
.action-item:hover {{ transform: translateY(-1px); }}
.action-header {{ display: flex; align-items: center; gap: 10px; flex-wrap: wrap; }}
.action-num {{ font-weight: 700; color: var(--cp-text-soft); font-size: 12px; }}
.severity-badge {{
    color: var(--cp-accent-fg); padding: 3px 10px; border-radius: 980px;
    font-size: 10px; font-weight: 700; text-transform: uppercase;
}}
.pillar-tag {{
    background: var(--cp-surface-soft); padding: 3px 10px; border-radius: 980px;
    font-size: 10px; color: var(--cp-text-muted);
}}
.action-title {{ font-weight: 590; font-size: 14px; }}
.action-desc {{ color: var(--cp-text-muted); margin-top: 8px; font-size: 13px; }}
.resource-list {{
    margin-top: 8px; padding-left: 20px;
    font-size: 11px; color: var(--cp-text-muted); max-height: 120px; overflow-y: auto;
}}
.resource-list li {{ margin: 2px 0; }}

/* Tables */
.data-grid {{
    display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-bottom: 24px;
}}
@media (max-width: 900px) {{ .data-grid {{ grid-template-columns: 1fr; }} }}
.table-card {{
    background: var(--cp-surface); border-radius: var(--cp-radius-md); padding: 22px;
    border: 1px solid var(--cp-border); box-shadow: var(--cp-shadow); min-width: 0;
}}
.table-card h3 {{ font-size: 14px; font-weight: 590; margin-bottom: 12px; color: var(--cp-text); }}
.table-card table {{ table-layout: fixed; }}
.table-card th,
.table-card td {{ overflow-wrap: anywhere; white-space: normal; }}
table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
th {{ background: transparent; padding: 8px; text-align: left; font-weight: 590; color: var(--cp-text-muted); border-bottom: 1px solid var(--cp-border-strong); }}
td {{ padding: 8px; border-bottom: 1px solid var(--cp-border); }}
tr:hover td {{ background: var(--cp-accent-soft); }}
.num {{ text-align: right; font-weight: 590; }}

/* Footer */
.footer {{
    text-align: center; padding: 32px 24px; color: var(--cp-text-soft); font-size: 11px;
    border-top: 1px solid var(--cp-border); margin-top: 32px;
}}

/* Header controls and bilingual toggle (default: English shown, Spanish hidden) */
.i18n-es {{ display: none; }}
body.lang-es .i18n-en {{ display: none; }}
body.lang-es .i18n-es {{ display: inline; }}
.header-controls {{
    position: absolute; top: 24px; right: 28px;
    display: flex; gap: 8px;
}}
.header-control {{
    background: rgba(255, 255, 255, 0.95); color: #1d1d1f;
    border: 1px solid rgba(255, 255, 255, 0.5); border-radius: 980px;
    height: 34px; min-width: 38px; padding: 0 14px;
    font-size: 12px; font-weight: 590; cursor: pointer;
    box-shadow: var(--cp-shadow); transition: transform 0.2s var(--cp-ease), background 0.2s var(--cp-ease);
}}
.header-control:hover {{ transform: translateY(-1px); background: #ffffff; }}
.theme-toggle {{ font-size: 18px; line-height: 1; padding: 0; }}
</style>
</head>
<body>
<div class="dashboard">

<!-- Header -->
<div class="header">
    <div class="header-controls">
        <button id="themeToggleBtn" class="header-control theme-toggle" onclick="toggleTheme()" type="button" aria-label="Switch to night mode" title="Switch to night mode">☾</button>
        <button id="langToggleBtn" class="header-control lang-toggle" onclick="toggleLang()" type="button">ES 🇪🇸</button>
    </div>
    <div class="header-brand">
        {self._azure_icon("azure-logo", "az-icon header-logo")}
        <h1>{bi("Azure WAF/CAF Workshop", "Taller Azure WAF/CAF")}<span class="header-title-sub">{bi("Discovery Report", "Informe de Descubrimiento")}</span></h1>
    </div>
    <div class="subtitle">{bi("Generated", "Generado")}: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | {bi("Consolidated view across all discovery phases", "Vista consolidada de todas las fases de descubrimiento")}</div>
</div>

<div class="view-tabs" id="viewTabsBar" role="tablist" aria-label="Dashboard views">
    <span id="viewTabThumb" class="view-tab-thumb" aria-hidden="true"></span>
    <button id="overviewTab" class="view-tab" role="tab" aria-selected="true" aria-controls="overviewView" onclick="setDashboardView('overview')" type="button">{bi('Overview', 'Resumen')}</button>
    <button id="reliabilityTab" class="view-tab" role="tab" aria-selected="false" aria-controls="reliabilityView" onclick="setDashboardView('reliability')" type="button">{bi('Reliability', 'Confiabilidad')}</button>
    <button id="securityTab" class="view-tab" role="tab" aria-selected="false" aria-controls="securityView" onclick="setDashboardView('security')" type="button">{bi('Security', 'Seguridad')}</button>
    <button id="costTab" class="view-tab" role="tab" aria-selected="false" aria-controls="costView" onclick="setDashboardView('cost')" type="button">{bi('Cost Opt', 'Optim. Costos')}</button>
    <button id="operationalTab" class="view-tab" role="tab" aria-selected="false" aria-controls="operationalView" onclick="setDashboardView('operational')" type="button">{bi('OpEx', 'Exc. Operativa')}</button>
    <button id="performanceTab" class="view-tab" role="tab" aria-selected="false" aria-controls="performanceView" onclick="setDashboardView('performance')" type="button">{bi('Performance', 'Rendimiento')}</button>
    <button id="governanceTab" class="view-tab" role="tab" aria-selected="false" aria-controls="governanceView" onclick="setDashboardView('governance')" type="button">{bi('Governance', 'Gobernanza')}</button>
</div>

<main id="overviewView" class="dashboard-view" role="tabpanel" aria-labelledby="overviewTab">
<!-- Overall Score + Pillar Breakdown -->
<div class="score-section">
    <div class="overall-score">
        <div class="score-circle"></div>
        <div class="score-label">{bi("Overall WAF Score", "Puntaje General WAF")}</div>
    </div>
    <div class="pillars-grid">
        {pillar_cards}
    </div>
</div>

<!-- Stats Row -->
<div class="stats-row">
    <div class="stat-box"><div class="value">{total_subs}</div><div class="label">{bi("Subscriptions", "Suscripciones")}</div></div>
    <div class="stat-box"><div class="value">{total_resources:,}</div><div class="label">{bi("Total Resources", "Recursos Totales")}</div></div>
    <div class="stat-box"><div class="value">{total_actions}</div><div class="label">{bi("Action Items", "Elementos de Acci\u00f3n")}</div></div>
    <div class="stat-box"><div class="value" style="color:var(--cp-danger);">{critical_actions}</div><div class="label">{bi("Critical/High Priority", "Prioridad Cr\u00edtica/Alta")}</div></div>
</div>

<!-- Score Methodology -->
<div class="section">
    <div class="section-title">📐 {bi("Score Methodology (Microsoft Advisor Score model)", "Metodolog\u00eda de Puntuaci\u00f3n (modelo de Azure Advisor Score)")}</div>
    <div class="methodology-note">
        {bi("Pillar scores follow the official", "Los puntajes de cada pilar siguen la f\u00f3rmula oficial de")} <a href="https://learn.microsoft.com/en-us/azure/advisor/advisor-score#calculation-of-advisor-score" target="_blank">Azure Advisor Score</a>
        {bi(
            "formulas and subcategory weights instead of an arbitrary point system. Security uses the Microsoft Defender Secure Score directly. "
            "Reliability, Performance, and Operational Excellence use Microsoft\u2019s published subcategory weights \u2014 recommendations are mapped to "
            "subcategories via keyword matching (Azure Resource Graph doesn\u2019t expose Advisor\u2019s internal subcategory tag), and the discovered "
            "resource inventory is used as the shared 'total applicable resources' pool. Cost uses a resource-count healthy ratio instead of "
            "retail-cost weighting, since this toolkit doesn\u2019t call the Azure Retail Prices API.",
            "y los pesos de subcategor\u00eda en lugar de un sistema de puntos arbitrario. Seguridad usa directamente el Secure Score de Microsoft Defender. "
            "Confiabilidad, Rendimiento y Excelencia Operativa usan los pesos de subcategor\u00eda publicados por Microsoft: las recomendaciones se "
            "asignan a subcategor\u00edas mediante coincidencia de palabras clave (Azure Resource Graph no expone la etiqueta interna de subcategor\u00eda de Advisor), "
            "y el inventario de recursos descubiertos se usa como el conjunto compartido de 'recursos totales aplicables'. Costo usa una proporci\u00f3n de "
            "recursos saludables en lugar de ponderaci\u00f3n por costo de lista, ya que esta herramienta no llama a la API de precios de Azure."
        )}
    </div>
    <div class="breakdown-grid">
        {score_breakdown_html}
    </div>
</div>

<!-- Action Items -->
<div class="section">
    <div class="section-title">🎯 {bi(f"Action Items ({total_actions})", f"Elementos de Acci\u00f3n ({total_actions})")}</div>
    {actions_html}
</div>

<!-- Data Summary Tables -->
<div class="data-grid">
    <div class="table-card">
        <h3>📦 {bi("Top Resource Types", "Principales Tipos de Recursos")}</h3>
        <table>
            <thead><tr><th>{bi("Resource Type", "Tipo de Recurso")}</th><th>{bi("Location", "Ubicaci\u00f3n")}</th><th>{bi("Count", "Cantidad")}</th></tr></thead>
            <tbody>{resource_rows}</tbody>
        </table>
    </div>
    <div class="table-card">
        <h3>💡 {bi("Advisor Summary", "Resumen de Advisor")}</h3>
        <table>
            <thead><tr><th>{bi("Category", "Categor\u00eda")}</th><th>{bi("Impact", "Impacto")}</th><th>{bi("Count", "Cantidad")}</th></tr></thead>
            <tbody>{advisor_rows}</tbody>
        </table>
    </div>
</div>
</main>

<main id="securityView" class="dashboard-view" role="tabpanel" aria-labelledby="securityTab" hidden>
    {security_view_html}
</main>

<main id="governanceView" class="dashboard-view" role="tabpanel" aria-labelledby="governanceTab" hidden>
    {governance_view_html}
</main>

<main id="costView" class="dashboard-view" role="tabpanel" aria-labelledby="costTab" hidden>
    {cost_view_html}
</main>

<main id="reliabilityView" class="dashboard-view" role="tabpanel" aria-labelledby="reliabilityTab" hidden>
    {reliability_view_html}
</main>

<main id="performanceView" class="dashboard-view" role="tabpanel" aria-labelledby="performanceTab" hidden>
    {performance_view_html}
</main>

<main id="operationalView" class="dashboard-view" role="tabpanel" aria-labelledby="operationalTab" hidden>
    {operational_view_html}
</main>

<!-- Footer -->
<div class="footer">
    {bi("Azure WAF/CAF Workshop Discovery Report | Generated by Azure Governance Discovery Toolkit", "Informe de Descubrimiento del Taller Azure WAF/CAF | Generado por Azure Governance Discovery Toolkit")}<br>
    {bi("This report consolidates findings from Resource Discovery, Azure Advisor, Metrics Analysis, and Governance Visualization.", "Este informe consolida los hallazgos de Resource Discovery, Azure Advisor, An\u00e1lisis de M\u00e9tricas y Visualizaci\u00f3n de Gobernanza.")}
</div>

</div>
<script>
function updateThemeButton() {{
    var btn = document.getElementById('themeToggleBtn');
    if (!btn) return;
    var dark = document.documentElement.getAttribute('data-theme') === 'dark';
    var spanish = document.body.classList.contains('lang-es');
    var label = dark
        ? (spanish ? 'Cambiar a modo claro' : 'Switch to light mode')
        : (spanish ? 'Cambiar a modo nocturno' : 'Switch to night mode');
    btn.textContent = dark ? '☀' : '☾';
    btn.setAttribute('aria-label', label);
    btn.setAttribute('title', label);
    btn.setAttribute('aria-pressed', String(dark));
}}
function setTheme(theme, persist) {{
    var normalized = theme === 'dark' ? 'dark' : 'light';
    document.documentElement.setAttribute('data-theme', normalized);
    if (persist !== false) {{
        try {{ localStorage.setItem('wafDashboardTheme', normalized); }} catch (e) {{}}
    }}
    updateThemeButton();
}}
function toggleTheme() {{
    setTheme(document.documentElement.getAttribute('data-theme') === 'dark' ? 'light' : 'dark');
}}
function setLang(lang) {{
    document.body.classList.toggle('lang-es', lang === 'es');
    document.documentElement.lang = lang;
    var btn = document.getElementById('langToggleBtn');
    if (btn) {{ btn.textContent = lang === 'es' ? 'EN 🇬🇧' : 'ES 🇪🇸'; }}
    updateThemeButton();
    try {{ localStorage.setItem('wafDashboardLang', lang); }} catch (e) {{}}
    moveTabThumb(currentDashboardView, true);
}}
function toggleLang() {{
    setLang(document.body.classList.contains('lang-es') ? 'en' : 'es');
}}
function toggleTableRows(btn) {{
    var wrap = btn.previousElementSibling;
    if (!wrap) return;
    var expanded = btn.classList.toggle('expanded');
    var rows = wrap.querySelectorAll('.extra-row');
    for (var i = 0; i < rows.length; i++) {{ rows[i].hidden = !expanded; }}
}}
var DASHBOARD_VIEWS = ['overview', 'reliability', 'security', 'cost', 'operational', 'performance', 'governance'];
var currentDashboardView = 'overview';
function moveTabThumb(view, skipAnimation) {{
    var thumb = document.getElementById('viewTabThumb');
    var tab = document.getElementById(view + 'Tab');
    var bar = document.getElementById('viewTabsBar');
    if (!thumb || !tab || !bar) return;
    if (skipAnimation) {{ thumb.style.transition = 'none'; }}
    var barRect = bar.getBoundingClientRect();
    var tabRect = tab.getBoundingClientRect();
    thumb.style.width = tabRect.width + 'px';
    thumb.style.transform = 'translateX(' + (tabRect.left - barRect.left) + 'px)';
    if (skipAnimation) {{
        void thumb.offsetHeight;
        thumb.style.transition = '';
    }}
}}
function setDashboardView(view, animate) {{
    var normalized = DASHBOARD_VIEWS.indexOf(view) !== -1 ? view : 'overview';
    animate = animate !== false;
    for (var i = 0; i < DASHBOARD_VIEWS.length; i++) {{
        var tab = document.getElementById(DASHBOARD_VIEWS[i] + 'Tab');
        if (tab) {{ tab.setAttribute('aria-selected', String(DASHBOARD_VIEWS[i] === normalized)); }}
    }}
    var prevPanel = document.getElementById(currentDashboardView + 'View');
    var nextPanel = document.getElementById(normalized + 'View');
    if (!animate || !prevPanel || prevPanel === nextPanel || !nextPanel) {{
        for (var j = 0; j < DASHBOARD_VIEWS.length; j++) {{
            var panel = document.getElementById(DASHBOARD_VIEWS[j] + 'View');
            if (panel) {{ panel.hidden = DASHBOARD_VIEWS[j] !== normalized; }}
        }}
    }} else {{
        prevPanel.classList.add('view-fade-out');
        window.setTimeout(function() {{
            prevPanel.hidden = true;
            prevPanel.classList.remove('view-fade-out');
            nextPanel.hidden = false;
            nextPanel.classList.add('view-fade-in');
            requestAnimationFrame(function() {{
                requestAnimationFrame(function() {{ nextPanel.classList.remove('view-fade-in'); }});
            }});
        }}, 180);
    }}
    currentDashboardView = normalized;
    moveTabThumb(normalized, !animate);
    try {{ localStorage.setItem('wafDashboardView', normalized); }} catch (e) {{}}
}}
window.addEventListener('resize', function() {{ moveTabThumb(currentDashboardView, true); }});
(function() {{
    setTheme(document.documentElement.getAttribute('data-theme') || 'light', false);
    var saved = 'en';
    try {{ saved = localStorage.getItem('wafDashboardLang') || 'en'; }} catch (e) {{}}
    setLang(saved);
    var savedView = 'overview';
    try {{ savedView = localStorage.getItem('wafDashboardView') || 'overview'; }} catch (e) {{}}
    setDashboardView(savedView, false);
}})();
</script>
</body>
</html>"""
        
        return html


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 generate-dashboard.py <output_directory>")
        print("  Where <output_directory> is the folder from Launch-AzureWorkshop.ps1")
        sys.exit(1)
    
    base_dir = sys.argv[1]
    if not os.path.isdir(base_dir):
        print(f"❌ Directory not found: {base_dir}")
        sys.exit(1)
    
    print(f"\n{'═'*60}")
    print(f"  Azure WAF/CAF Workshop - Dashboard Generator")
    print(f"{'═'*60}")
    print(f"  Input: {base_dir}\n")
    
    # Read all data
    print("📂 Reading data files...")
    data = ExcelReader.read_all(base_dir)
    
    # Analyze
    print("\n🔍 Analyzing across WAF pillars...")
    engine = InsightEngine(data).analyze()
    
    print(f"\n📊 Scores:")
    for pillar, score in engine.scores.items():
        bar = "█" * (score // 5) + "░" * (20 - score // 5)
        status = "✅" if score >= 80 else "⚠️ " if score >= 50 else "🔴"
        print(f"   {status} {pillar:<25} {bar} {score}/100")
    print(f"\n   Overall WAF Score: {engine.get_overall_score()}/100")
    print(f"   Action Items: {len(engine.action_items)}")
    
    # Generate dashboard
    print("\n🎨 Generating dashboard HTML...")
    generator = DashboardGenerator(engine, data, base_dir)
    html = generator.generate()
    
    output_path = os.path.join(base_dir, "07_Dashboard", "WAF_Dashboard.html")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    
    print(f"\n{'═'*60}")
    print(f"  ✅ Dashboard generated!")
    print(f"  📁 {output_path}")
    print(f"  💡 Open in any browser to view the interactive report.")
    print(f"{'═'*60}\n")


if __name__ == "__main__":
    main()
